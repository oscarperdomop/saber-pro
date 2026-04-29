import io
import re

import pandas as pd
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import CharField, Count, Q, Value
from django.db.models.functions import Coalesce, Concat
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import Notificacion, ProgramaAcademico, Usuario
from .serializers import (
    ActualizarMiPasswordSerializer,
    CambiarPasswordSerializer,
    CustomTokenObtainPairSerializer,
    MiPerfilSerializer,
    NotificacionSerializer,
    ProgramaAcademicoSerializer,
    UsuarioCreateSerializer,
    UsuarioListadoSerializer,
    validate_new_password_rules,
)

# Acepta letras ASCII y latin extendido (tildes/dieresis/enye), mas espacios.
SOLO_LETRAS_REGEX = re.compile(r'^[A-Za-z\u00C0-\u00FF\s]+$')


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class ActivarCuentaView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        serializer = CambiarPasswordSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        password_actual = serializer.validated_data['password_actual']
        password_nueva = serializer.validated_data['password_nueva']

        if not request.user.check_password(password_actual):
            return Response(
                {'password_actual': ['La contrasena actual es incorrecta.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        request.user.set_password(password_nueva)
        request.user.es_primer_ingreso = False
        request.user.save()

        return Response(
            {'mensaje': 'Contrasena actualizada y cuenta activada con exito.'},
            status=status.HTTP_200_OK,
        )


class CargaMasivaUsuariosView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    parser_classes = (MultiPartParser, FormParser)

    @staticmethod
    def _normalizar_mojibake(value):
        """
        Corrige textos dañados por una decodificacion incorrecta UTF-8/Latin-1:
        ejemplo: PÃ©rez -> Pérez, IngenierÃ­a -> Ingeniería.
        """
        if not isinstance(value, str) or not value:
            return value

        if '\u00C3' not in value and '\u00C2' not in value:
            return value

        for source in ('latin-1', 'cp1252'):
            try:
                fixed = value.encode(source).decode('utf-8')
                if ('\u00C3' not in fixed and '\u00C2' not in fixed) or fixed.count('\uFFFD') < value.count('\uFFFD'):
                    return fixed
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
        return value

    @staticmethod
    def _clean_value(value):
        if value is None:
            return ''
        text = CargaMasivaUsuariosView._normalizar_mojibake(str(value))
        text = text.strip()
        if text.endswith('.0') and text.replace('.', '', 1).isdigit():
            text = text[:-2]
        return text

    @staticmethod
    def _limit_length(value, max_len):
        return value[:max_len]

    @staticmethod
    def _to_upper(value):
        return value.upper() if isinstance(value, str) else value

    @staticmethod
    def _read_cell(row, *keys):
        for key in keys:
            value = row.get(key, None)
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
        return ''

    @staticmethod
    def _parse_genero(value):
        normalized = str(value or '').strip().upper()
        if not normalized:
            return None
        if normalized.startswith('M'):
            return 'M'
        if normalized.startswith('F'):
            return 'F'
        if normalized.startswith('O'):
            return 'O'
        return None

    @staticmethod
    def _parse_semestre(value):
        normalized = str(value or '').strip()
        if not normalized:
            return None
        try:
            semestre = int(float(normalized))
        except (TypeError, ValueError):
            return None
        return semestre if semestre > 0 else None

    @staticmethod
    def _read_csv_uploaded_file(uploaded_file):
        """
        Lee CSV priorizando UTF-8 con BOM (utf-8-sig), que es el formato recomendado
        para archivos abiertos en Excel en Windows.
        """
        raw_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        if not raw_bytes:
            raise ValueError('El archivo CSV esta vacio.')

        # Camino principal: UTF-8 con BOM. Evita que Excel/Python dejen caracteres raros.
        try:
            utf8_sig_text = raw_bytes.decode('utf-8-sig')
            if '\x00' not in utf8_sig_text:
                return pd.read_csv(io.StringIO(utf8_sig_text), sep=None, engine='python')
        except UnicodeDecodeError:
            pass

        # Fallback defensivo para archivos legados.
        last_error = None
        for encoding in ('utf-8', 'utf-16', 'cp1252', 'latin-1'):
            try:
                text = raw_bytes.decode(encoding)
                if '\x00' in text:
                    continue
                text = CargaMasivaUsuariosView._normalizar_mojibake(text)
                return pd.read_csv(io.StringIO(text), sep=None, engine='python')
            except UnicodeDecodeError as exc:
                last_error = exc
                continue

        raise ValueError(f'No se pudo decodificar el archivo CSV: {last_error}')

    def post(self, request):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        archivo = request.FILES.get('archivo') or request.FILES.get('file')
        if not archivo:
            return Response(
                {'detalle': "Debes enviar un archivo en el campo 'archivo' o 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        nombre_archivo = archivo.name.lower()

        try:
            if nombre_archivo.endswith('.xlsx'):
                df = pd.read_excel(archivo)
            elif nombre_archivo.endswith('.csv'):
                # Detecta delimitador automaticamente (coma, punto y coma, tab, etc.)
                # usando stream de texto, no bytes.
                df = self._read_csv_uploaded_file(archivo)
            else:
                return Response(
                    {'detalle': 'Formato de archivo no soportado. Usa .xlsx o .csv.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as exc:
            return Response(
                {'detalle': f'No se pudo leer el archivo de carga: {exc}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        df.fillna('', inplace=True)

        creados = 0
        errores = []
        documentos_existentes = {
            str(value).strip().upper()
            for value in Usuario.objects.values_list('documento', flat=True)
            if value
        }
        correos_existentes = {
            str(value).strip().lower()
            for value in Usuario.objects.values_list('correo_institucional', flat=True)
            if value
        }
        documentos_archivo = set()
        correos_archivo = set()

        for index, row in df.iterrows():
            try:
                with transaction.atomic():
                    documento = self._limit_length(
                        self._clean_value(self._read_cell(row, 'documento', 'DOCUMENTO')),
                        20,
                    )
                    tipo_documento = self._limit_length(
                        self._clean_value(self._read_cell(row, 'tipo_documento', 'TIPO_DOCUMENTO')),
                        5,
                    )
                    nombres = self._limit_length(
                        self._clean_value(self._read_cell(row, 'nombres', 'NOMBRES')),
                        100,
                    )
                    apellidos = self._limit_length(
                        self._clean_value(self._read_cell(row, 'apellidos', 'APELLIDOS')),
                        100,
                    )
                    correo_institucional = self._limit_length(
                        self._clean_value(
                            self._read_cell(row, 'correo_institucional', 'CORREO_INSTITUCIONAL'),
                        ).lower(),
                        254,
                    )
                    nombre_programa = self._limit_length(
                        self._clean_value(self._read_cell(row, 'programa', 'PROGRAMA')),
                        150,
                    )
                    genero_raw = self._clean_value(
                        self._read_cell(row, 'genero', 'GENERO', 'género', 'GÉNERO'),
                    )
                    semestre_raw = self._clean_value(
                        self._read_cell(row, 'semestre_actual', 'semestre', 'SEMESTRE'),
                    )
                    genero = self._parse_genero(genero_raw)
                    semestre_actual = self._parse_semestre(semestre_raw)

                    documento = self._to_upper(documento)
                    tipo_documento = self._to_upper(tipo_documento)
                    nombres = self._to_upper(nombres)
                    apellidos = self._to_upper(apellidos)
                    nombre_programa = self._to_upper(nombre_programa)

                    if not all(
                        [
                            documento,
                            tipo_documento,
                            nombres,
                            apellidos,
                            correo_institucional,
                            nombre_programa,
                        ]
                        ):
                        raise ValueError(
                            'Faltan campos obligatorios: documento, tipo_documento, nombres, '
                            'apellidos, correo_institucional, programa.'
                        )

                    nombres = ' '.join(nombres.split())
                    apellidos = ' '.join(apellidos.split())

                    if not SOLO_LETRAS_REGEX.fullmatch(nombres):
                        raise ValueError('Nombres invalidos: solo se permiten letras.')
                    if not SOLO_LETRAS_REGEX.fullmatch(apellidos):
                        raise ValueError('Apellidos invalidos: solo se permiten letras.')

                    if not documento.isdigit():
                        raise ValueError('Documento invalido: solo se permiten numeros.')
                    if len(documento) > 10:
                        raise ValueError(
                            'Documento invalido: debe tener maximo 10 digitos.',
                        )

                    if not correo_institucional.endswith('@usco.edu.co'):
                        raise ValueError(
                            'Correo institucional invalido: debe terminar en @usco.edu.co.',
                        )

                    if documento in documentos_existentes:
                        raise ValueError('Documento duplicado: ya existe en la base de datos.')
                    if correo_institucional in correos_existentes:
                        raise ValueError(
                            'Correo institucional duplicado: ya existe en la base de datos.',
                        )
                    if documento in documentos_archivo:
                        raise ValueError('Documento duplicado dentro del archivo de carga.')
                    if correo_institucional in correos_archivo:
                        raise ValueError(
                            'Correo institucional duplicado dentro del archivo de carga.',
                        )

                    codigo_snies = self._clean_value(row.get('codigo_snies')) or documento
                    codigo_snies = self._limit_length(codigo_snies, 15)
                    facultad = self._limit_length(self._clean_value(row.get('facultad')) or 'SIN DEFINIR', 100)
                    sede = self._limit_length(self._clean_value(row.get('sede')) or 'PRINCIPAL', 50)

                    codigo_snies = self._to_upper(codigo_snies)
                    facultad = self._to_upper(facultad)
                    sede = self._to_upper(sede)

                    programa_obj, _ = ProgramaAcademico.objects.get_or_create(
                        nombre=nombre_programa,
                        defaults={
                            'codigo_snies': codigo_snies,
                            'facultad': facultad,
                            'sede': sede,
                        },
                    )

                    Usuario.objects.create_user(
                        correo_institucional=correo_institucional,
                        password=documento,
                        documento=documento,
                        tipo_documento=tipo_documento,
                        nombres=nombres,
                        apellidos=apellidos,
                        programa=programa_obj,
                        genero=genero,
                        semestre_actual=semestre_actual,
                        es_primer_ingreso=True,
                    )

                    documentos_existentes.add(documento)
                    correos_existentes.add(correo_institucional)
                    documentos_archivo.add(documento)
                    correos_archivo.add(correo_institucional)
                    creados += 1
            except Exception as e:
                errores.append({'fila': index + 2, 'error': str(e)})

        return Response(
            {'total_procesados': len(df), 'creados': creados, 'errores': errores},
            status=status.HTTP_200_OK,
        )


class UsuariosListView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    search_fields = (
        'nombres',
        'apellidos',
        'correo_institucional',
        'documento',
        'nombre_completo',
    )

    class UsuariosPagination(PageNumberPagination):
        page_size = 10
        page_size_query_param = 'page_size'
        max_page_size = 50

    def get_queryset(self):
        return (
            Usuario.objects.select_related('programa')
            .annotate(
                nombre_completo=Concat(
                    Coalesce('nombres', Value('')),
                    Value(' '),
                    Coalesce('apellidos', Value('')),
                    output_field=CharField(),
                ),
            )
            .all()
            .order_by('nombres', 'apellidos')
        )

    def _apply_search_filters(self, queryset, search):
        normalized_search = ' '.join((search or '').split())
        if not normalized_search:
            return queryset

        search_query = Q()
        for field in self.search_fields:
            search_query |= Q(**{f'{field}__icontains': normalized_search})

        return queryset.filter(search_query)

    @staticmethod
    def _normalize_bool_query(value):
        normalized = str(value or '').strip().lower()
        if normalized in {'true', '1', 'activo', 'activos'}:
            return True
        if normalized in {'false', '0', 'inactivo', 'inactivos'}:
            return False
        return None

    def _apply_request_filters(self, queryset, request):
        search = (request.query_params.get('search') or '').strip()
        rol = str(request.query_params.get('rol') or '').strip().upper()
        estado = request.query_params.get('is_active')
        # Compatibilidad hacia atras con el parametro previo "estado".
        if estado in (None, ''):
            estado = request.query_params.get('estado')
        programa = str(request.query_params.get('programa') or '').strip()
        semestre = str(request.query_params.get('semestre') or '').strip()

        queryset = self._apply_search_filters(queryset, search)

        if rol in {Usuario.ROL_ADMIN, Usuario.ROL_PROFESOR, Usuario.ROL_ESTUDIANTE}:
            queryset = queryset.filter(rol=rol)

        estado_bool = self._normalize_bool_query(estado)
        if estado_bool is not None:
            queryset = queryset.filter(is_active=estado_bool)

        if programa:
            if programa.isdigit():
                queryset = queryset.filter(programa_id=int(programa))
            else:
                queryset = queryset.filter(programa__nombre__icontains=programa)

        if semestre:
            try:
                queryset = queryset.filter(semestre_actual=int(semestre))
            except ValueError:
                pass

        return queryset

    def get(self, request):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        usuarios = self._apply_request_filters(self.get_queryset(), request)

        paginator = self.UsuariosPagination()
        page = paginator.paginate_queryset(usuarios, request, view=self)
        serializer = UsuarioListadoSerializer(page, many=True)

        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = UsuarioCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        usuario = serializer.save(es_primer_ingreso=True, is_active=True)
        serializer = UsuarioListadoSerializer(usuario)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class UsuariosExportExcelView(UsuariosListView):
    permission_classes = [IsAdminUser]

    @staticmethod
    def _estado_label(is_active):
        return 'Activo' if bool(is_active) else 'Inactivo'

    def get(self, request):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        usuarios = self._apply_request_filters(self.get_queryset(), request)
        rows = usuarios.values_list(
            'documento',
            'nombres',
            'apellidos',
            'correo_institucional',
            'programa__nombre',
            'rol',
            'is_active',
            'semestre_actual',
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Usuarios'
        worksheet.freeze_panes = 'A2'

        # Estilo homologado con reportes de resultados de simulacros.
        header_fill = PatternFill(fill_type='solid', fgColor='8F141B')
        header_font = Font(color='FFFFFF', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')

        headers = [
            'Documento',
            'Nombres',
            'Apellidos',
            'Correo',
            'Programa',
            'Rol',
            'Estado',
            'Semestre',
        ]
        worksheet.append(headers)
        for column_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for (
            documento,
            nombres,
            apellidos,
            correo,
            programa_nombre,
            rol,
            is_active,
            semestre_actual,
        ) in rows:
            worksheet.append(
                [
                    documento or '',
                    nombres or '',
                    apellidos or '',
                    correo or '',
                    programa_nombre or 'Sin programa',
                    rol or '',
                    self._estado_label(is_active),
                    semestre_actual if semestre_actual is not None else '',
                ]
            )
            row_index = worksheet.max_row
            worksheet.cell(row=row_index, column=1).alignment = center_alignment
            worksheet.cell(row=row_index, column=2).alignment = left_alignment
            worksheet.cell(row=row_index, column=3).alignment = left_alignment
            worksheet.cell(row=row_index, column=4).alignment = left_alignment
            worksheet.cell(row=row_index, column=5).alignment = left_alignment
            worksheet.cell(row=row_index, column=6).alignment = center_alignment
            worksheet.cell(row=row_index, column=7).alignment = center_alignment
            worksheet.cell(row=row_index, column=8).alignment = center_alignment

        for index, header in enumerate(headers, start=1):
            max_length = len(header)
            col_values = next(
                worksheet.iter_cols(
                    min_col=index,
                    max_col=index,
                    min_row=2,
                    max_row=worksheet.max_row,
                    values_only=True,
                ),
                [],
            )
            for value in col_values:
                cell_value = '' if value is None else str(value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(max_length + 2, 12),
                50,
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=\"reporte_usuarios.xlsx\"'
        return response


class UsuariosDashboardStatsView(APIView):
    permission_classes = [IsAdminUser]

    @staticmethod
    def _build_roles_payload(rows):
        counters = {'ADMIN': 0, 'PROFESOR': 0, 'ESTUDIANTE': 0}
        for row in rows:
            rol = str(row.get('rol') or '').strip().upper()
            total = int(row.get('total') or 0)
            if rol in counters:
                counters[rol] = total

        return {
            'admin': counters['ADMIN'],
            'profesor': counters['PROFESOR'],
            'estudiante': counters['ESTUDIANTE'],
            'detalle': [
                {'rol': 'ADMIN', 'total': counters['ADMIN']},
                {'rol': 'PROFESOR', 'total': counters['PROFESOR']},
                {'rol': 'ESTUDIANTE', 'total': counters['ESTUDIANTE']},
            ],
        }

    @staticmethod
    def _build_genero_payload(rows):
        map_labels = {'M': 'Masculino', 'F': 'Femenino', 'O': 'Otro', 'SIN_DATO': 'Sin especificar'}
        payload = []
        for row in rows:
            key = str(row.get('genero_key') or 'SIN_DATO')
            payload.append(
                {
                    'genero': key,
                    'genero_nombre': map_labels.get(key, key),
                    'total': int(row.get('total') or 0),
                }
            )
        return payload

    @staticmethod
    def _build_semestre_payload(rows):
        payload = []
        for row in rows:
            semestre = row.get('semestre_key')
            payload.append(
                {
                    'semestre': int(semestre) if semestre is not None else None,
                    'semestre_nombre': f'Semestre {int(semestre)}' if semestre is not None else 'Sin dato',
                    'total': int(row.get('total') or 0),
                }
            )
        return payload

    def get(self, request):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        usuarios_qs = Usuario.objects.select_related('programa').all()

        total_usuarios = usuarios_qs.count()
        usuarios_activos = usuarios_qs.filter(is_active=True).count()

        roles_rows = usuarios_qs.values('rol').annotate(total=Count('id')).order_by('rol')
        por_rol = self._build_roles_payload(roles_rows)

        genero_rows = (
            usuarios_qs.annotate(genero_key=Coalesce('genero', Value('SIN_DATO')))
            .values('genero_key')
            .annotate(total=Count('id'))
            .order_by('genero_key')
        )
        por_genero = self._build_genero_payload(genero_rows)

        programa_rows = (
            usuarios_qs.annotate(programa_nombre=Coalesce('programa__nombre', Value('Sin programa')))
            .values('programa_nombre')
            .annotate(total=Count('id'))
            .order_by('-total', 'programa_nombre')
        )
        por_programa = [
            {
                'programa': str(row.get('programa_nombre') or 'Sin programa'),
                'total': int(row.get('total') or 0),
            }
            for row in programa_rows
        ]

        semestre_rows = (
            usuarios_qs.annotate(semestre_key=Coalesce('semestre_actual', Value(None)))
            .values('semestre_key')
            .annotate(total=Count('id'))
            .order_by('semestre_key')
        )
        por_semestre = self._build_semestre_payload(semestre_rows)

        return Response(
            {
                'total_usuarios': total_usuarios,
                'usuarios_activos': usuarios_activos,
                'total_estudiantes': por_rol.get('estudiante', 0),
                'por_rol': por_rol,
                'por_genero': por_genero,
                'por_programa': por_programa,
                'por_semestre': por_semestre,
            },
            status=status.HTTP_200_OK,
        )


class ProgramasListView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        programas = ProgramaAcademico.objects.all().order_by('nombre')
        serializer = ProgramaAcademicoSerializer(programas, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class MiPerfilView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = MiPerfilSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request):
        serializer = ActualizarMiPasswordSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        password_nueva = serializer.validated_data['password_nueva']
        request.user.set_password(password_nueva)
        request.user.save(update_fields=['password'])

        return Response({'mensaje': 'Contrasena actualizada con exito.'}, status=status.HTTP_200_OK)


class NotificacionesListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        limite_raw = str(request.query_params.get('limit') or '').strip()
        try:
            limite = int(limite_raw) if limite_raw else 20
        except ValueError:
            limite = 20

        limite = max(1, min(limite, 100))
        notificaciones = Notificacion.objects.filter(usuario=request.user).order_by('-created_at')[:limite]
        serializer = NotificacionSerializer(notificaciones, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request):
        ids = request.data.get('ids')
        marcar_todas_raw = request.data.get('marcar_todas')
        if isinstance(marcar_todas_raw, bool):
            marcar_todas = marcar_todas_raw
        else:
            marcar_todas = str(marcar_todas_raw or '').strip().lower() in {'1', 'true', 'yes', 'si'}

        queryset = Notificacion.objects.filter(usuario=request.user, leida=False)
        if not marcar_todas:
            if not isinstance(ids, list) or len(ids) == 0:
                return Response(
                    {'detalle': "Envia 'ids' (lista) o 'marcar_todas': true."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            queryset = queryset.filter(id__in=ids)

        actualizadas = queryset.update(leida=True)
        return Response({'actualizadas': actualizadas}, status=status.HTTP_200_OK)


class NotificacionesContadorView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        no_leidas = Notificacion.objects.filter(usuario=request.user, leida=False).count()
        return Response({'no_leidas': no_leidas}, status=status.HTTP_200_OK)


class UsuariosEstadoUpdateView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    @staticmethod
    def _parse_bool(value):
        if isinstance(value, bool):
            return value

        if isinstance(value, str):
            normalized = value.strip().lower()
            if normalized in {'true', '1', 'yes', 'si'}:
                return True
            if normalized in {'false', '0', 'no'}:
                return False

        return None

    @staticmethod
    def _parse_genero(value):
        normalized = str(value or '').strip().upper()
        if not normalized:
            return None
        if normalized.startswith('M'):
            return 'M'
        if normalized.startswith('F'):
            return 'F'
        if normalized.startswith('O'):
            return 'O'
        return None

    @staticmethod
    def _parse_semestre(value):
        normalized = str(value or '').strip()
        if not normalized:
            return None
        try:
            semestre = int(float(normalized))
        except (TypeError, ValueError):
            return None
        return semestre if semestre > 0 else None

    def patch(self, request, user_id):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        usuario = get_object_or_404(Usuario, id=user_id)
        payload = request.data
        campos_perfil = {
            'nombres',
            'apellidos',
            'correo_institucional',
            'tipo_documento',
            'numero_documento',
            'documento',
            'rol',
            'is_staff',
            'programa_id',
            'programa',
            'genero',
            'semestre_actual',
        }
        password_nueva = str(payload.get('password') or '').strip()
        requiere_actualizacion_estado = 'is_active' in payload
        requiere_actualizacion_perfil = any(campo in payload for campo in campos_perfil)
        requiere_actualizacion_password = bool(password_nueva)

        if (
            not requiere_actualizacion_estado
            and not requiere_actualizacion_perfil
            and not requiere_actualizacion_password
        ):
            return Response(
                {'detalle': 'No se enviaron campos validos para actualizar.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        update_fields = []

        if requiere_actualizacion_estado:
            is_active = self._parse_bool(payload.get('is_active'))
            if is_active is None:
                return Response(
                    {'detalle': "El campo 'is_active' es obligatorio y debe ser booleano."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            usuario.is_active = is_active
            update_fields.append('is_active')

        if requiere_actualizacion_perfil:
            if 'nombres' in payload:
                nombres = ' '.join(str(payload.get('nombres') or '').strip().split())
                if not nombres or not all(char.isalpha() or char.isspace() for char in nombres):
                    return Response(
                        {'nombres': ['Los nombres solo deben contener letras.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.nombres = nombres.upper()
                update_fields.append('nombres')

            if 'apellidos' in payload:
                apellidos = ' '.join(str(payload.get('apellidos') or '').strip().split())
                if not apellidos or not all(char.isalpha() or char.isspace() for char in apellidos):
                    return Response(
                        {'apellidos': ['Los apellidos solo deben contener letras.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.apellidos = apellidos.upper()
                update_fields.append('apellidos')

            if 'correo_institucional' in payload:
                correo = str(payload.get('correo_institucional') or '').strip().lower()
                if not correo.endswith('@usco.edu.co'):
                    return Response(
                        {
                            'correo_institucional': [
                                'El correo institucional debe terminar en @usco.edu.co.',
                            ],
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if Usuario.objects.filter(correo_institucional__iexact=correo).exclude(id=usuario.id).exists():
                    return Response(
                        {'correo_institucional': ['Ya existe un usuario con este correo institucional.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.correo_institucional = correo
                update_fields.append('correo_institucional')

            if 'tipo_documento' in payload:
                usuario.tipo_documento = str(payload.get('tipo_documento') or '').strip().upper()
                update_fields.append('tipo_documento')

            if 'numero_documento' in payload or 'documento' in payload:
                documento = payload.get('numero_documento', payload.get('documento'))
                documento_text = str(documento or '').strip()
                if not documento_text.isdigit():
                    return Response(
                        {'numero_documento': ['El numero de documento solo debe contener numeros.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if len(documento_text) > 10:
                    return Response(
                        {'numero_documento': ['El numero de documento debe tener maximo 10 digitos.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if Usuario.objects.filter(documento__iexact=documento_text).exclude(id=usuario.id).exists():
                    return Response(
                        {'numero_documento': ['Ya existe un usuario con este numero de documento.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.documento = documento_text
                update_fields.append('documento')

            if 'rol' in payload:
                rol = str(payload.get('rol') or '').strip().upper()
                if rol not in {Usuario.ROL_ADMIN, Usuario.ROL_PROFESOR, Usuario.ROL_ESTUDIANTE}:
                    return Response(
                        {
                            'rol': [
                                (
                                    'Rol invalido. Usa '
                                    f'{Usuario.ROL_ADMIN}, {Usuario.ROL_PROFESOR} o {Usuario.ROL_ESTUDIANTE}.'
                                ),
                            ],
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.rol = rol
                update_fields.append('rol')

            if 'is_staff' in payload:
                is_staff = self._parse_bool(payload.get('is_staff'))
                if is_staff is None:
                    return Response(
                        {'is_staff': ["El campo 'is_staff' debe ser booleano."]},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                rol_objetivo = (
                    str(payload.get('rol') or '').strip().upper() if 'rol' in payload else usuario.rol
                )
                if rol_objetivo == Usuario.ROL_PROFESOR:
                    usuario.is_staff = is_staff
                elif rol_objetivo == Usuario.ROL_ADMIN:
                    usuario.is_staff = True
                else:
                    usuario.is_staff = False
                update_fields.append('is_staff')

            if 'programa_id' in payload or 'programa' in payload:
                programa_id = payload.get('programa_id', payload.get('programa'))
                if programa_id in (None, ''):
                    usuario.programa = None
                else:
                    try:
                        programa_id_int = int(programa_id)
                    except (TypeError, ValueError):
                        return Response(
                            {'programa_id': ['Programa invalido.']},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    usuario.programa = get_object_or_404(ProgramaAcademico, id=programa_id_int)
                update_fields.append('programa')

            if 'genero' in payload:
                genero = self._parse_genero(payload.get('genero'))
                valor_genero = str(payload.get('genero') or '').strip()
                if valor_genero and genero is None:
                    return Response(
                        {'genero': ['Genero invalido. Usa M, F u O.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.genero = genero
                update_fields.append('genero')

            if 'semestre_actual' in payload:
                semestre = self._parse_semestre(payload.get('semestre_actual'))
                valor_semestre = str(payload.get('semestre_actual') or '').strip()
                if valor_semestre and semestre is None:
                    return Response(
                        {'semestre_actual': ['Semestre invalido. Debe ser numerico y mayor a 0.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                usuario.semestre_actual = semestre
                update_fields.append('semestre_actual')

            if usuario.rol != Usuario.ROL_ESTUDIANTE:
                if usuario.semestre_actual is not None:
                    usuario.semestre_actual = None
                    update_fields.append('semestre_actual')
            elif usuario.semestre_actual is None:
                return Response(
                    {
                        'semestre_actual': [
                            'El semestre actual es obligatorio para usuarios con rol ESTUDIANTE.',
                        ],
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if requiere_actualizacion_password:
            try:
                validate_new_password_rules(usuario, password_nueva)
            except (ValidationError, DRFValidationError) as exc:
                mensajes = []
                if hasattr(exc, 'message_dict'):
                    for values in exc.message_dict.values():
                        mensajes.extend(values)
                if hasattr(exc, 'messages'):
                    mensajes.extend(exc.messages)
                if not mensajes:
                    mensajes.append(str(exc))

                return Response({'password': mensajes}, status=status.HTTP_400_BAD_REQUEST)

            usuario.set_password(password_nueva)
            update_fields.append('password')

        try:
            usuario.full_clean()
        except ValidationError as exc:
            return Response(exc.message_dict, status=status.HTTP_400_BAD_REQUEST)

        usuario.save(update_fields=list(dict.fromkeys(update_fields)))

        serializer = UsuarioListadoSerializer(usuario)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, user_id):
        if getattr(request.user, 'rol', None) != Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'Solo el Administrador Supremo puede gestionar usuarios.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        usuario = get_object_or_404(Usuario, id=user_id)

        if str(usuario.id) == str(request.user.id):
            return Response(
                {'detalle': 'No puedes eliminar tu propio usuario.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if usuario.is_superuser or usuario.rol == Usuario.ROL_ADMIN:
            return Response(
                {'detalle': 'No se puede eliminar el superusuario.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tiene_intentos = usuario.intentos.exists()

        if tiene_intentos:
            if usuario.is_active:
                usuario.is_active = False
                usuario.save(update_fields=['is_active'])
                mensaje = (
                    'El usuario tiene intentos de examen registrados; '
                    'se aplico eliminacion logica (desactivado).'
                )
            else:
                mensaje = (
                    'El usuario tiene intentos de examen y ya estaba inactivo '
                    '(eliminacion logica ya aplicada).'
                )

            return Response(
                {
                    'status': 'OK',
                    'tipo_eliminacion': 'logica',
                    'usuario_id': str(usuario.id),
                    'mensaje': mensaje,
                },
                status=status.HTTP_200_OK,
            )

        usuario_id = str(usuario.id)
        usuario.delete()
        return Response(
            {
                'status': 'OK',
                'tipo_eliminacion': 'fisica',
                'usuario_id': usuario_id,
                'mensaje': 'Usuario eliminado permanentemente de la base de datos.',
            },
            status=status.HTTP_200_OK,
        )
