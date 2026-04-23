import csv
import json
import re
import time
import uuid
from io import BytesIO, StringIO
import unicodedata
from decimal import Decimal

import google.generativeai as genai
import pandas as pd
from django.conf import settings
from django.db import transaction
from django.db.models import Avg, Case, Count, DurationField, ExpressionWrapper, F, FloatField, Q, Value, When
from django.db.models.functions import Cast
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.utils import timezone
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from core.permissions import IsOwnerOrSupremo
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    Categoria,
    Competencia,
    IntentoExamen,
    IntentoModulo,
    ModuloPrueba,
    OpcionRespuesta,
    PlantillaExamen,
    Pregunta,
    RespuestaEstudiante,
)
from users.models import ProgramaAcademico, Usuario
from .utils import compilar_preview_latex_base64, obtener_distractores_ia
from .serializers import (
    CategoriaAdminSerializer,
    CalificarEnsayoSerializer,
    CompetenciaAdminSerializer,
    EnsayoPendienteSerializer,
    IntentoExamenSerializer,
    ModuloSerializer,
    PlantillaExamenAdminSerializer,
    PlantillaExamenEstudianteSerializer,
    PreguntaAdminSerializer,
    RevisionRespuestaSerializer,
    RespuestaEstudianteDetalleSerializer,
    RespuestaEstudianteUpdateSerializer,
)


class GenerarOpcionesIAView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    @staticmethod
    def _clean_json_response(texto):
        contenido = (texto or '').strip()

        if contenido.startswith('```json'):
            contenido = contenido.removeprefix('```json').strip()
        elif contenido.startswith('```'):
            contenido = contenido.removeprefix('```').strip()

        if contenido.endswith('```'):
            contenido = contenido.removesuffix('```').strip()

        return contenido

    @staticmethod
    def _resolver_modelo_gemini():
        modelos_disponibles = []
        for model in genai.list_models():
            supported = list(getattr(model, 'supported_generation_methods', []) or [])
            if 'generateContent' in supported:
                modelos_disponibles.append(getattr(model, 'name', ''))

        if not modelos_disponibles:
            raise ValueError('No hay modelos Gemini con soporte generateContent en esta cuenta/API key.')

        override_model = str(getattr(settings, 'GEMINI_MODEL', '') or '').strip()
        if override_model:
            candidate = override_model if override_model.startswith('models/') else f'models/{override_model}'
            if candidate in modelos_disponibles:
                return candidate, modelos_disponibles

        preferidos = [
            'models/gemini-2.5-flash',
            'models/gemini-2.0-flash',
            'models/gemini-2.0-flash-lite',
            'models/gemini-1.5-flash-latest',
            'models/gemini-1.5-flash',
        ]

        for candidato in preferidos:
            if candidato in modelos_disponibles:
                return candidato, modelos_disponibles

        return modelos_disponibles[0], modelos_disponibles

    def post(self, request):
        modelos_disponibles = []
        try:
            genai.configure(api_key=settings.GEMINI_API_KEY)

            enunciado = str(request.data.get('enunciado') or '').strip()
            contexto = str(request.data.get('contexto') or '').strip()

            if not enunciado:
                return Response(
                    {'detalle': 'El campo enunciado es obligatorio para generar opciones con IA.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            prompt_text = f"""
Eres un experto pedagogo creador de exámenes tipo ICFES Saber Pro en Colombia.
Dado el siguiente enunciado y contexto, genera 4 opciones de respuesta.
Solo 1 debe ser correcta, y las otras 3 deben ser distractores plausibles que evalúen errores comunes del estudiante.
Devuelve el resultado ESTRICTAMENTE como un array de objetos JSON con esta estructura exacta, sin texto adicional, sin formato markdown:
[
    {{"texto": "opción correcta aquí", "es_correcta": true}},
    {{"texto": "distractor 1", "es_correcta": false}},
    {{"texto": "distractor 2", "es_correcta": false}},
    {{"texto": "distractor 3", "es_correcta": false}}
]

Contexto: {contexto if contexto else 'Ninguno'}
Enunciado: {enunciado}
"""

            modelo_nombre, modelos_disponibles = self._resolver_modelo_gemini()
            model = genai.GenerativeModel(modelo_nombre)
            response = model.generate_content(prompt_text)
            raw_text = getattr(response, 'text', '') or ''

            cleaned_text = self._clean_json_response(raw_text)
            opciones = json.loads(cleaned_text)

            if not isinstance(opciones, list):
                raise ValueError('La IA no devolvio un array JSON valido.')

            opciones_normalizadas = []
            for opcion in opciones:
                if not isinstance(opcion, dict):
                    continue
                opciones_normalizadas.append(
                    {
                        'texto': str(opcion.get('texto') or '').strip(),
                        'es_correcta': bool(opcion.get('es_correcta', False)),
                    }
                )

            if not opciones_normalizadas:
                raise ValueError('La IA devolvio opciones vacias o con formato invalido.')

            return Response(opciones_normalizadas, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'detalle': 'No fue posible generar opciones con IA.',
                    'error': str(e),
                    'sugerencia_modelo': 'Configura GEMINI_MODEL en backend/.env con un modelo disponible.',
                    'modelos_disponibles': modelos_disponibles[:10],
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class LaTeXPreviewView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        texto_latex = str(request.data.get('texto_latex') or '').strip()
        if not texto_latex:
            return Response(
                {
                    'status': 'error',
                    'detalle': 'El campo texto_latex es obligatorio.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        timeout_seconds = int(getattr(settings, 'LATEX_PREVIEW_TIMEOUT_SECONDS', 5) or 5)
        timeout_seconds = max(3, min(timeout_seconds, 5))
        pdf_base64, png_base64, compile_error = compilar_preview_latex_base64(
            texto_latex,
            timeout_seconds=timeout_seconds,
        )

        if compile_error:
            return Response(
                {
                    'status': 'error',
                    'detalle': 'No fue posible compilar el codigo LaTeX.',
                    'detalle_tecnico': compile_error,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                'status': 'ok',
                'pdf_base64': pdf_base64,
                'preview_png_base64': png_base64,
            },
            status=status.HTTP_200_OK,
        )


class PreguntaAdminViewSet(viewsets.ModelViewSet):
    queryset = Pregunta.objects.all().order_by('created_at', 'id')
    serializer_class = PreguntaAdminSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    _option_pattern = re.compile(r'^opciones\[(\d+)\]\[(id|texto|es_correcta|imagen)\]$')

    def get_queryset(self):
        queryset = Pregunta.objects.all().order_by('created_at', 'id')
        incluir_archivadas = str(
            self.request.query_params.get('incluir_archivadas')
            or self.request.query_params.get('include_archivadas')
            or ''
        ).strip().lower() in ['1', 'true', 'yes']
        modulo_id = str(self.request.query_params.get('modulo_id') or '').strip()
        modulo_nombre = str(self.request.query_params.get('modulo_nombre') or '').strip()
        estado = str(self.request.query_params.get('estado') or '').strip()
        estado_normalizado = estado.lower()
        tipo_pregunta = str(self.request.query_params.get('tipo_pregunta') or '').strip()
        dificultad = str(self.request.query_params.get('dificultad') or '').strip()
        ordering = str(self.request.query_params.get('ordering') or '').strip()

        # "Todas/Todos" debe significar sin filtro de estado e incluir archivadas.
        if estado_normalizado in ['todas', 'todos']:
            estado = ''
            incluir_archivadas = True

        if self.action == 'list' and not incluir_archivadas:
            queryset = queryset.exclude(estado='Archivada')

        if modulo_id:
            queryset = queryset.filter(modulo_id=modulo_id)

        if modulo_nombre:
            queryset = queryset.filter(modulo__nombre__iexact=modulo_nombre)

        if estado:
            queryset = queryset.filter(estado__iexact=estado)

        if tipo_pregunta:
            queryset = queryset.filter(tipo_pregunta=tipo_pregunta)

        if dificultad:
            nivel_map = {
                'facil': 'Facil',
                'fácil': 'Facil',
                'media': 'Medio',
                'medio': 'Medio',
                'alta': 'Dificil',
                'dificil': 'Dificil',
                'difícil': 'Dificil',
            }
            dificultad_normalizada = nivel_map.get(dificultad.lower())
            if dificultad_normalizada:
                queryset = queryset.filter(nivel_dificultad=dificultad_normalizada)

        allowed_ordering = {'created_at', '-created_at', 'enunciado', '-enunciado'}
        if ordering in allowed_ordering:
            queryset = queryset.order_by(ordering, 'id')

        return queryset

    @staticmethod
    def _map_dificultad(dificultad):
        return {
            'Facil': 'Facil',
            'Media': 'Medio',
            'Alta': 'Dificil',
            'Medio': 'Medio',
            'Dificil': 'Dificil',
        }.get(str(dificultad or '').strip(), 'Medio')

    @staticmethod
    def _map_estado_masivo(estado):
        estado_limpio = str(estado or '').strip().lower()
        estado_map = {
            'publicada': 'Publicada',
            'activo': 'Publicada',
            'activa': 'Publicada',
            'borrador': 'Borrador',
            'inactivo': 'Borrador',
            'inactiva': 'Borrador',
            'archivada': 'Archivada',
        }
        return estado_map.get(estado_limpio)

    @staticmethod
    def _to_bool(value):
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ['1', 'true', 'yes', 'on']

    def _extract_opciones_formdata(self, request):
        grouped = {}
        found = False

        for key in request.data.keys():
            match = self._option_pattern.match(str(key))
            if not match:
                continue
            found = True
            index = int(match.group(1))
            field = match.group(2)
            grouped.setdefault(index, {})
            grouped[index][field] = request.data.get(key)

        for key in request.FILES.keys():
            match = self._option_pattern.match(str(key))
            if not match:
                continue
            found = True
            index = int(match.group(1))
            field = match.group(2)
            grouped.setdefault(index, {})
            grouped[index][field] = request.FILES.get(key)

        if not found:
            return None

        opciones = []
        for index in sorted(grouped.keys()):
            data = grouped[index]
            opcion_id = str(data.get('id') or '').strip() or None
            texto = str(data.get('texto') or '').strip()
            es_correcta = self._to_bool(data.get('es_correcta'))
            imagen = data.get('imagen')

            if not texto and not imagen and not opcion_id:
                continue

            opciones.append(
                {
                    'id': opcion_id,
                    'texto': texto or None,
                    'es_correcta': es_correcta,
                    'imagen': imagen,
                }
            )

        return opciones

    @staticmethod
    def _extract_base_data(request):
        allowed_fields = [
            'modulo',
            'modulo_id',
            'categoria',
            'competencia',
            'nivel_dificultad',
            'dificultad',
            'tipo_pregunta',
            'contexto_texto',
            'codigo_latex',
            'soporte_multimedia',
            'enunciado',
            'justificacion',
            'limite_palabras',
            'rubrica_evaluacion',
            'estado',
            'version_original',
        ]
        data = {}
        for field in allowed_fields:
            if field in request.data:
                data[field] = request.data.get(field)

        if 'contexto_imagen' in request.FILES:
            data['contexto_imagen'] = request.FILES.get('contexto_imagen')
        if 'imagen_grafica' in request.FILES:
            data['imagen_grafica'] = request.FILES.get('imagen_grafica')

        return data

    @staticmethod
    def _validate_manual_options(tipo_pregunta, opciones):
        if tipo_pregunta == 'Ensayo':
            return None

        if len(opciones) < 2:
            return Response(
                {'opciones': ['Debes agregar al menos dos opciones para una pregunta de opcion multiple.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not any(opcion.get('es_correcta') for opcion in opciones):
            return Response(
                {'opciones': ['Debes marcar al menos una opcion correcta.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return None

    @staticmethod
    def _replace_opciones(pregunta, opciones):
        existing_by_id = {str(opcion.id): opcion for opcion in pregunta.opciones.all()}
        pregunta.opciones.all().delete()
        for opcion in opciones:
            opcion_id = str(opcion.get('id') or '').strip()
            imagen = opcion.get('imagen')
            if not imagen and opcion_id and opcion_id in existing_by_id:
                imagen = existing_by_id[opcion_id].imagen

            OpcionRespuesta.objects.create(
                pregunta=pregunta,
                texto=opcion.get('texto'),
                es_correcta=bool(opcion.get('es_correcta')),
                imagen=imagen,
            )

    @staticmethod
    def _hydrate_missing_option_images(opciones, source_pregunta):
        existing_by_id = {str(opcion.id): opcion for opcion in source_pregunta.opciones.all()}
        hydrated = []

        for opcion in opciones:
            opcion_copy = {**opcion}
            opcion_id = str(opcion_copy.get('id') or '').strip()
            if not opcion_copy.get('imagen') and opcion_id and opcion_id in existing_by_id:
                opcion_copy['imagen'] = existing_by_id[opcion_id].imagen
            hydrated.append(opcion_copy)

        return hydrated

    @staticmethod
    def _clone_opciones(instance):
        return [
            {
                'texto': opcion.texto,
                'es_correcta': opcion.es_correcta,
                'imagen': opcion.imagen if opcion.imagen else None,
            }
            for opcion in instance.opciones.all()
        ]

    @staticmethod
    def _clean_cell(value):
        if value is None:
            return ''
        try:
            if pd.isna(value):
                return ''
        except TypeError:
            pass
        text = str(value).strip()
        if text.lower() in ['nan', 'none', '<na>']:
            return ''
        if text.endswith('.0') and text.replace('.', '', 1).isdigit():
            return text[:-2]
        return text

    @staticmethod
    def _normalize_lookup_text(value):
        text = str(value or '').strip().lower()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(char for char in text if not unicodedata.combining(char))
        return ' '.join(text.split())

    @staticmethod
    def _normalize_column_name(value):
        text = str(value or '').replace('\ufeff', '').strip()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(char for char in text if not unicodedata.combining(char))
        text = text.upper()
        text = re.sub(r'[^A-Z0-9]+', '_', text).strip('_')
        return text

    @staticmethod
    def _sanitize_filename_fragment(value):
        text = str(value or '').strip()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(char for char in text if not unicodedata.combining(char))
        text = re.sub(r'[^A-Za-z0-9]+', '_', text).strip('_')
        return text or 'generica'

    @classmethod
    def _read_cell(cls, row, *keys):
        for key in keys:
            normalized_key = cls._normalize_column_name(key)
            if normalized_key not in row:
                continue
            value = cls._clean_cell(row.get(normalized_key))
            if value:
                return value
        return ''

    @staticmethod
    def _resolve_modulo(raw_modulo):
        modulo_raw = str(raw_modulo or '').strip()
        if not modulo_raw:
            return None

        if modulo_raw.isdigit():
            return ModuloPrueba.objects.filter(id=int(modulo_raw)).first()

        modulo = ModuloPrueba.objects.filter(nombre__iexact=modulo_raw).first()
        if modulo:
            return modulo

        objetivo = PreguntaAdminViewSet._normalize_lookup_text(modulo_raw)
        for candidato in ModuloPrueba.objects.all():
            if PreguntaAdminViewSet._normalize_lookup_text(candidato.nombre) == objetivo:
                return candidato

        return None

    @staticmethod
    def _resolve_categoria(modulo, categoria_default, raw_categoria):
        categoria_raw = str(raw_categoria or '').strip()
        if not categoria_raw:
            return categoria_default

        if categoria_raw.isdigit():
            return Categoria.objects.filter(id=int(categoria_raw), modulo=modulo).first()

        categoria = Categoria.objects.filter(modulo=modulo, nombre__iexact=categoria_raw).first()
        if categoria:
            return categoria

        objetivo = PreguntaAdminViewSet._normalize_lookup_text(categoria_raw)
        for candidato in Categoria.objects.filter(modulo=modulo):
            if PreguntaAdminViewSet._normalize_lookup_text(candidato.nombre) == objetivo:
                return candidato

        return None

    @staticmethod
    def _resolve_competencia(modulo, competencia_default, raw_competencia):
        competencia_raw = str(raw_competencia or '').strip()
        if not competencia_raw:
            return competencia_default

        if competencia_raw.isdigit():
            return Competencia.objects.filter(id=int(competencia_raw), modulo=modulo).first()

        competencia = Competencia.objects.filter(modulo=modulo, nombre__iexact=competencia_raw).first()
        if competencia:
            return competencia

        objetivo = PreguntaAdminViewSet._normalize_lookup_text(competencia_raw)
        for candidato in Competencia.objects.filter(modulo=modulo):
            if PreguntaAdminViewSet._normalize_lookup_text(candidato.nombre) == objetivo:
                return candidato

        return None

    @classmethod
    def _read_bulk_file(cls, archivo):
        nombre_archivo = str(getattr(archivo, 'name', '') or '').lower()
        if nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls'):
            df = pd.read_excel(archivo)
        elif nombre_archivo.endswith('.csv'):
            archivo.seek(0)
            raw_content = archivo.read()

            if isinstance(raw_content, bytes):
                decoded = None
                for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
                    try:
                        decoded = raw_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                if decoded is None:
                    raise ValueError(
                        'No se pudo leer el CSV. Verifica que este guardado en UTF-8 o Latin-1.'
                    )
                csv_text = decoded
            else:
                csv_text = str(raw_content or '')

            lines = csv_text.splitlines()
            if lines and lines[0].strip().lower().startswith('sep='):
                lines = lines[1:]
            csv_text = '\n'.join(lines)

            sample_line = next((line for line in lines if line.strip()), '')
            if sample_line.count(';') >= sample_line.count(','):
                delimiter = ';'
            else:
                delimiter = ','

            df = pd.read_csv(StringIO(csv_text), sep=delimiter, engine='python')
        else:
            raise ValueError('Formato no soportado. Usa .xlsx, .xls o .csv.')

        # Normaliza nombres de columnas para tolerar BOM, tildes, espacios y variantes.
        df.columns = [cls._normalize_column_name(column) for column in df.columns]

        return df

    def create(self, request, *args, **kwargs):
        opciones_formdata = self._extract_opciones_formdata(request)
        if opciones_formdata is None:
            data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        else:
            data = self._extract_base_data(request)

        modulo_valor = data.get('modulo') or data.get('modulo_id')
        if modulo_valor and not data.get('modulo'):
            data['modulo'] = modulo_valor

        if data.get('dificultad') and not data.get('nivel_dificultad'):
            data['nivel_dificultad'] = self._map_dificultad(data.get('dificultad'))

        serializer = self.get_serializer(
            data=data,
            context={
                **self.get_serializer_context(),
                'skip_options_validation': opciones_formdata is not None,
            },
        )
        serializer.is_valid(raise_exception=True)
        tipo_pregunta = str(data.get('tipo_pregunta') or 'Opcion Multiple')

        if opciones_formdata is not None:
            validation_error = self._validate_manual_options(tipo_pregunta, opciones_formdata)
            if validation_error is not None:
                return validation_error

        self.perform_create(serializer)
        pregunta = serializer.instance

        if opciones_formdata is not None:
            if tipo_pregunta == 'Ensayo':
                pregunta.opciones.all().delete()
            else:
                self._replace_opciones(pregunta, opciones_formdata)

        response_data = self.get_serializer(pregunta).data
        headers = self.get_success_headers(response_data)
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        if instance.estado == 'Archivada':
            return Response(
                {'detalle': 'Las preguntas archivadas son de solo lectura y no pueden modificarse.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        opciones_formdata = self._extract_opciones_formdata(request)

        if opciones_formdata is None:
            data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        else:
            data = self._extract_base_data(request)

        modulo_valor = data.get('modulo') or data.get('modulo_id')
        if modulo_valor and not data.get('modulo'):
            data['modulo'] = modulo_valor

        if data.get('dificultad') and not data.get('nivel_dificultad'):
            data['nivel_dificultad'] = self._map_dificultad(data.get('dificultad'))

        serializer = self.get_serializer(
            instance,
            data=data,
            partial=partial,
            context={
                **self.get_serializer_context(),
                'skip_options_validation': opciones_formdata is not None,
            },
        )
        serializer.is_valid(raise_exception=True)

        tipo_pregunta = str(
            data.get('tipo_pregunta')
            or ('Ensayo' if instance.limite_palabras is not None else 'Opcion Multiple')
        )

        if opciones_formdata is not None:
            validation_error = self._validate_manual_options(tipo_pregunta, opciones_formdata)
            if validation_error is not None:
                return validation_error

        tiene_respuestas = RespuestaEstudiante.objects.filter(pregunta=instance).exists()
        if tiene_respuestas:
            with transaction.atomic():
                estado_previo = instance.estado
                instance.estado = 'Archivada'
                instance.save(update_fields=['estado', 'updated_at'])

                data_version = {
                    'modulo_id': data.get('modulo_id') or data.get('modulo') or instance.modulo_id,
                    'categoria': data.get('categoria', instance.categoria_id),
                    'competencia': data.get('competencia', instance.competencia_id),
                    'nivel_dificultad': data.get('nivel_dificultad', instance.nivel_dificultad),
                    'tipo_pregunta': data.get(
                        'tipo_pregunta',
                        'Ensayo' if instance.limite_palabras is not None else 'Opcion Multiple',
                    ),
                    'contexto_texto': (
                        data.get('contexto_texto') if 'contexto_texto' in data else instance.contexto_texto
                    ),
                    'codigo_latex': (
                        data.get('codigo_latex') if 'codigo_latex' in data else instance.codigo_latex
                    ),
                    'soporte_multimedia': (
                        data.get('soporte_multimedia')
                        if 'soporte_multimedia' in data
                        else instance.soporte_multimedia
                    ),
                    'enunciado': data.get('enunciado', instance.enunciado),
                    'justificacion': (
                        data.get('justificacion') if 'justificacion' in data else instance.justificacion
                    ),
                    'limite_palabras': (
                        data.get('limite_palabras') if 'limite_palabras' in data else instance.limite_palabras
                    ),
                    'rubrica_evaluacion': (
                        data.get('rubrica_evaluacion')
                        if 'rubrica_evaluacion' in data
                        else instance.rubrica_evaluacion
                    ),
                    'estado': data.get('estado', estado_previo),
                    'version_original': data.get('version_original')
                    or instance.version_original_id
                    or instance.id,
                }

                if 'contexto_imagen' in data:
                    data_version['contexto_imagen'] = data.get('contexto_imagen')
                elif instance.contexto_imagen:
                    data_version['contexto_imagen'] = instance.contexto_imagen
                if 'imagen_grafica' in data:
                    data_version['imagen_grafica'] = data.get('imagen_grafica')
                elif instance.imagen_grafica:
                    data_version['imagen_grafica'] = instance.imagen_grafica

                opciones_version = (
                    opciones_formdata if opciones_formdata is not None else self._clone_opciones(instance)
                )
                opciones_version = self._hydrate_missing_option_images(opciones_version, instance)

                validation_error = self._validate_manual_options(
                    str(data_version.get('tipo_pregunta') or 'Opcion Multiple'),
                    opciones_version,
                )
                if validation_error is not None:
                    return validation_error

                serializer_version = self.get_serializer(
                    data=data_version,
                    context={
                        **self.get_serializer_context(),
                        'skip_options_validation': True,
                    },
                )
                serializer_version.is_valid(raise_exception=True)
                serializer_version.save()
                pregunta_nueva = serializer_version.instance

                if str(data_version.get('tipo_pregunta')) == 'Ensayo':
                    pregunta_nueva.opciones.all().delete()
                else:
                    self._replace_opciones(pregunta_nueva, opciones_version)

                return Response(
                    {
                        'mensaje': (
                            'Se creó una nueva versión porque la original ya tenía respuestas.'
                        ),
                        'versionada': True,
                        'nueva_pregunta': self.get_serializer(pregunta_nueva).data,
                    },
                    status=status.HTTP_201_CREATED,
                )

        self.perform_update(serializer)
        pregunta = serializer.instance

        if opciones_formdata is not None:
            if tipo_pregunta == 'Ensayo':
                pregunta.opciones.all().delete()
            else:
                self._replace_opciones(pregunta, opciones_formdata)

        return Response(self.get_serializer(pregunta).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        pregunta_id = str(instance.id)
        tiene_respuestas = RespuestaEstudiante.objects.filter(pregunta=instance).exists()

        if tiene_respuestas:
            if instance.estado != 'Archivada':
                instance.estado = 'Archivada'
                instance.save(update_fields=['estado', 'updated_at'])
                mensaje = (
                    'La pregunta tiene respuestas asociadas y fue archivada '
                    'para conservar la trazabilidad historica.'
                )
            else:
                mensaje = (
                    'La pregunta ya se encontraba archivada porque tiene '
                    'respuestas asociadas.'
                )

            return Response(
                {
                    'status': 'OK',
                    'tipo_eliminacion': 'logica',
                    'pregunta_id': pregunta_id,
                    'mensaje': mensaje,
                },
                status=status.HTTP_200_OK,
            )

        self.perform_destroy(instance)
        return Response(
            {
                'status': 'OK',
                'tipo_eliminacion': 'fisica',
                'pregunta_id': pregunta_id,
                'mensaje': 'Pregunta eliminada permanentemente de la base de datos.',
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=['patch'],
        url_path='bulk-update-estado',
        permission_classes=[IsAuthenticated, IsAdminUser],
    )
    def bulk_update_estado(self, request):
        ids = request.data.get('ids')
        estado_destino = self._map_estado_masivo(request.data.get('estado'))

        if not isinstance(ids, list) or not ids:
            return Response(
                {'detalle': 'Debes enviar una lista no vacia en el campo ids.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ids_unicos = []
        vistos = set()
        for value in ids:
            raw_id = str(value or '').strip()
            if not raw_id or raw_id in vistos:
                continue
            ids_unicos.append(raw_id)
            vistos.add(raw_id)

        if not ids_unicos:
            return Response(
                {'detalle': 'No se recibieron IDs validos para actualizar.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not estado_destino:
            return Response(
                {
                    'detalle': (
                        "Estado invalido. Usa uno de: 'Publicada', 'Borrador' o 'Archivada'."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(id__in=ids_unicos)
        encontrados = queryset.count()
        no_encontrados = max(len(ids_unicos) - encontrados, 0)

        if encontrados == 0:
            return Response(
                {'detalle': 'No se encontraron preguntas para los IDs enviados.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        bloqueadas_historial = queryset.filter(estado='Archivada').count()
        queryset_editable = queryset.exclude(estado='Archivada')
        sin_cambio = queryset_editable.filter(estado=estado_destino).count()
        actualizados = queryset_editable.exclude(estado=estado_destino).update(
            estado=estado_destino,
            updated_at=timezone.now(),
        )

        return Response(
            {
                'status': 'ok',
                'actualizados': actualizados,
                'sin_cambio': sin_cambio,
                'bloqueadas_historial': bloqueadas_historial,
                'no_encontrados': no_encontrados,
                'estado_aplicado': estado_destino,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'], url_path='plantilla-carga')
    def plantilla_carga(self, request):
        modulo_id = str(request.query_params.get('modulo_id') or '').strip()
        modulo_nombre = str(request.query_params.get('modulo_nombre') or '').strip()
        modulo_param = str(request.query_params.get('modulo') or '').strip()

        modulo_seleccionado = None

        if modulo_id:
            if not modulo_id.isdigit():
                return Response(
                    {'detalle': 'El parametro modulo_id es invalido.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            modulo_seleccionado = ModuloPrueba.objects.filter(id=int(modulo_id)).first()
            if modulo_seleccionado is None:
                return Response(
                    {'detalle': 'No existe un modulo con el modulo_id enviado.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            modulo_lookup = modulo_nombre or modulo_param
            if modulo_lookup:
                modulo_seleccionado = self._resolve_modulo(modulo_lookup)
                if modulo_seleccionado is None:
                    return Response(
                        {'detalle': 'No existe un modulo con el nombre enviado.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        modulo_fijo = str(getattr(modulo_seleccionado, 'nombre', '') or '').strip()
        categorias_modulo = []
        competencias_modulo = []

        if modulo_seleccionado is not None:
            categorias_modulo = list(
                Categoria.objects.filter(modulo=modulo_seleccionado)
                .order_by('nombre')
                .values_list('nombre', flat=True)
            )
            competencias_modulo = list(
                Competencia.objects.filter(modulo=modulo_seleccionado)
                .order_by('nombre')
                .values_list('nombre', flat=True)
            )

        rows = [
            [
                'Razonamiento Cuantitativo',
                'Resuelve: x^2 - x - 6 = 0. Cual es una raiz?',
                'Algebra y calculo',
                'Interpretacion y representacion',
                '3',
                '',
                '',
                '',
                'Media',
                'Publicada',
            ],
            [
                'Razonamiento Cuantitativo',
                'En un triangulo rectangulo, calcula la hipotenusa.',
                'Geometria',
                'Resolucion de problemas',
                '10',
                '8',
                '9',
                '12',
                'Media',
                'Borrador',
            ],
            [
                'Razonamiento Cuantitativo',
                'El promedio de 4, 6 y 8 es:',
                'Estadistica',
                'Formulacion y ejecucion',
                '6',
                '5',
                '7',
                '8',
                'Facil',
                'Publicada',
            ],
            [
                'Lectura Critica',
                'Segun el texto, cual es la idea principal?',
                'Lectura Critica',
                'Comprension e interpretacion textual',
                'La idea central del autor',
                '',
                '',
                '',
                'Media',
                'Publicada',
            ],
        ]

        if modulo_fijo:
            if categorias_modulo:
                rows_dinamicas = []
                for index, categoria_nombre in enumerate(categorias_modulo):
                    competencia_derivada = (
                        competencias_modulo[index % len(competencias_modulo)]
                        if competencias_modulo
                        else ''
                    )
                    rows_dinamicas.append(
                        [
                            modulo_fijo,
                            f'Ejemplo de pregunta para la categoria {categoria_nombre}.',
                            categoria_nombre,
                            competencia_derivada,
                            'Respuesta correcta de ejemplo',
                            'Distractor 1 de ejemplo',
                            'Distractor 2 de ejemplo',
                            'Distractor 3 de ejemplo',
                            'Media',
                            'Borrador',
                        ]
                    )
                rows = rows_dinamicas
            else:
                rows = [[modulo_fijo, *row[1:]] for row in rows]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Plantilla Preguntas'

        headers = [
            'MODULO',
            'ENUNCIADO',
            'CATEGORIA',
            'COMPETENCIA',
            'OPCION_CORRECTA',
            'DISTRACTOR_1',
            'DISTRACTOR_2',
            'DISTRACTOR_3',
            'DIFICULTAD',
            'ESTADO',
        ]

        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill(start_color='FF7A0019', end_color='FF7A0019', fill_type='solid')
        header_font = Font(color='FFFFFFFF', bold=True)
        predefinido_fill = PatternFill(start_color='FFFDE68A', end_color='FFFDE68A', fill_type='solid')

        for column_index, _ in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if modulo_fijo:
            for row_index in range(2, len(rows) + 2):
                modulo_cell = worksheet.cell(row=row_index, column=1)
                modulo_cell.value = modulo_fijo
                modulo_cell.fill = predefinido_fill
                modulo_cell.font = Font(bold=True, color='FF854D0E')

        worksheet.freeze_panes = 'A2'
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 58
        worksheet.column_dimensions['C'].width = 26
        worksheet.column_dimensions['D'].width = 34
        worksheet.column_dimensions['E'].width = 18
        worksheet.column_dimensions['F'].width = 18
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 18
        worksheet.column_dimensions['I'].width = 14
        worksheet.column_dimensions['J'].width = 14

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename_base = (
            f"plantilla_preguntas_{self._sanitize_filename_fragment(modulo_fijo)}"
            if modulo_fijo
            else 'plantilla_preguntas_generica'
        )

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    @action(
        detail=False,
        methods=['post'],
        parser_classes=[MultiPartParser, FormParser],
        url_path='carga-masiva',
    )
    def carga_masiva(self, request):
        archivo = request.FILES.get('archivo') or request.FILES.get('file')
        if not archivo:
            return Response(
                {'detalle': "Debes enviar un archivo en el campo 'archivo' o 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        modulo_id = str(request.data.get('modulo_id') or '').strip()
        modulo_forzado = None
        if modulo_id:
            try:
                modulo_forzado = ModuloPrueba.objects.get(id=int(modulo_id))
            except (TypeError, ValueError, ModuloPrueba.DoesNotExist):
                return Response(
                    {'detalle': 'Modulo invalido.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        categoria_forzada = None
        competencia_forzada = None
        categoria_id = str(request.data.get('categoria_id') or '').strip()
        competencia_id = str(request.data.get('competencia_id') or '').strip()

        if (categoria_id or competencia_id) and modulo_forzado is None:
            return Response(
                {
                    'detalle': (
                        'Para sobrescribir categoria o competencia debes seleccionar '
                        'un modulo global en el modal.'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if categoria_id:
            categoria_forzada = Categoria.objects.filter(
                id=categoria_id,
                modulo=modulo_forzado,
            ).first()
            if categoria_forzada is None:
                return Response(
                    {'detalle': 'La categoria seleccionada no pertenece al modulo indicado.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if competencia_id:
            competencia_forzada = Competencia.objects.filter(
                id=competencia_id,
                modulo=modulo_forzado,
            ).first()
            if competencia_forzada is None:
                return Response(
                    {'detalle': 'La competencia seleccionada no pertenece al modulo indicado.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        usar_ia_param = str(request.data.get('usar_ia', 'true')).strip().lower() == 'true'

        try:
            df = self._read_bulk_file(archivo)
        except Exception as exc:
            return Response({'detalle': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        estado_default = str(request.data.get('estado') or 'Publicada').strip() or 'Publicada'
        dificultad_default = self._map_dificultad(request.data.get('dificultad') or 'Media')
        nuevo_lote = uuid.uuid4()

        creadas = 0
        omitidas = 0
        filas_con_ia = 0
        sin_ia_por_error = 0
        filas_omitidas = 0
        errores = []

        for index, row in df.iterrows():
            fila_excel = index + 2
            try:
                with transaction.atomic():
                    enunciado = self._read_cell(row, 'ENUNCIADO')
                    opcion_correcta = self._read_cell(
                        row,
                        'OPCION_CORRECTA',
                        'RESPUESTA_CORRECTA',
                    )
                    modulo_row = self._read_cell(row, 'MODULO_ID', 'MODULO')
                    categoria_row = self._read_cell(row, 'CATEGORIA_ID', 'CATEGORIA')
                    competencia_row = self._read_cell(
                        row,
                        'COMPETENCIA_ID',
                        'COMPETENCIA',
                    )
                    distractor_1 = self._read_cell(row, 'DISTRACTOR_1')
                    distractor_2 = self._read_cell(row, 'DISTRACTOR_2')
                    distractor_3 = self._read_cell(row, 'DISTRACTOR_3')

                    # Ignora filas totalmente vacias (comunes al final de archivos CSV/XLSX).
                    if not any(
                        [
                            enunciado,
                            opcion_correcta,
                            modulo_row,
                            categoria_row,
                            competencia_row,
                            distractor_1,
                            distractor_2,
                            distractor_3,
                        ]
                    ):
                        filas_omitidas += 1
                        continue

                    modulo_final = modulo_forzado
                    if modulo_final is None:
                        modulo_final = self._resolve_modulo(modulo_row)

                    if modulo_final is None:
                        raise ValueError(
                            'No se pudo resolver el modulo para la fila. '
                            'Completa MODULO/MODULO_ID en el Excel o selecciona un modulo en el modal.'
                        )

                    if not enunciado or not opcion_correcta:
                        raise ValueError(
                            'Cada fila debe incluir ENUNCIADO y OPCION_CORRECTA.'
                        )

                    enunciado_limpio = str(enunciado).strip()
                    existe_duplicado = Pregunta.objects.filter(
                        enunciado__iexact=enunciado_limpio,
                        modulo=modulo_final,
                    ).exists()

                    if existe_duplicado:
                        omitidas += 1
                        continue

                    contexto_texto = self._read_cell(
                        row, 'CONTEXTO', 'contexto', 'CONTEXTO_TEXTO', 'contexto_texto'
                    )
                    estado = self._read_cell(row, 'ESTADO', 'estado') or estado_default
                    dificultad_val = (
                        self._read_cell(
                            row,
                            'DIFICULTAD',
                            'dificultad',
                            'NIVEL_DIFICULTAD',
                            'nivel_dificultad',
                        )
                        or dificultad_default
                    )
                    nivel_dificultad = self._map_dificultad(dificultad_val)

                    categoria = self._resolve_categoria(
                        modulo_final,
                        None,
                        categoria_row,
                    )
                    competencia = self._resolve_competencia(
                        modulo_final,
                        None,
                        competencia_row,
                    )

                    if categoria_forzada is not None:
                        categoria = categoria_forzada

                    if competencia_forzada is not None:
                        competencia = competencia_forzada

                    if not categoria:
                        raise ValueError(
                            'No se pudo resolver la categoria para la fila. '
                            'Completa CATEGORIA/CATEGORIA_ID en el Excel o selecciona una categoria en el modal.'
                        )
                    if not competencia:
                        raise ValueError(
                            'No se pudo resolver la competencia para la fila. '
                            'Completa COMPETENCIA/COMPETENCIA_ID en el Excel o selecciona una competencia en el modal.'
                        )

                    distractores_raw = [
                        distractor_1,
                        distractor_2,
                        distractor_3,
                    ]
                    distractores = [value for value in distractores_raw if value]

                    if len(distractores) < 3 and usar_ia_param:
                        longitud_inicial_distractores = len(distractores)
                        try:
                            sugeridos = obtener_distractores_ia(
                                enunciado=enunciado_limpio,
                                opcion_correcta=opcion_correcta,
                                modulo=str(getattr(modulo_final, 'nombre', '') or '').strip() or 'General',
                                categoria=str(getattr(categoria, 'nombre', '') or '').strip() or 'General',
                            )
                            # Throttling para respetar cuotas de la API gratuita.
                            time.sleep(5)
                        except Exception as exc:
                            sin_ia_por_error += 1
                            errores.append(
                                {
                                    'fila': fila_excel,
                                    'error': (
                                        "Error o cuota excedida en IA. "
                                        f"Se guarda la pregunta sin distractores generados. Detalle: {str(exc)}"
                                    ),
                                }
                            )
                            sugeridos = []

                        for sugerido in sugeridos:
                            limpio = str(sugerido or '').strip()
                            if not limpio:
                                continue
                            if limpio.lower() == opcion_correcta.lower():
                                continue
                            if limpio.lower() in {d.lower() for d in distractores}:
                                continue
                            distractores.append(limpio)
                            if len(distractores) == 3:
                                break

                        if len(distractores) > longitud_inicial_distractores:
                            filas_con_ia += 1

                    if len(distractores) < 3:
                        while len(distractores) < 3:
                            distractores.append('')

                    pregunta = Pregunta.objects.create(
                        modulo=modulo_final,
                        categoria=categoria,
                        competencia=competencia,
                        nivel_dificultad=nivel_dificultad,
                        contexto_texto=contexto_texto or None,
                        enunciado=enunciado_limpio,
                        estado=estado,
                        lote_id=nuevo_lote,
                        limite_palabras=None,
                    )

                    OpcionRespuesta.objects.create(
                        pregunta=pregunta,
                        texto=opcion_correcta,
                        es_correcta=True,
                    )
                    for distractor in distractores[:3]:
                        OpcionRespuesta.objects.create(
                            pregunta=pregunta,
                            texto=distractor,
                            es_correcta=False,
                        )

                    creadas += 1
            except Exception as exc:
                errores.append({'fila': fila_excel, 'error': str(exc)})

        return Response(
            {
                'status': 'OK',
                'mensaje': (
                    f'Proceso finalizado. Creadas: {creadas}. '
                    f'Omitidas por duplicado: {omitidas}.'
                ),
                'creadas': creadas,
                'omitidas': omitidas,
                'preguntas_creadas': creadas,
                'filas_con_ia': filas_con_ia,
                'sin_ia_por_error': sin_ia_por_error,
                'filas_omitidas': filas_omitidas,
                'lote_id': str(nuevo_lote),
                'errores': errores,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['post'], url_path='revertir-carga')
    def revertir_carga(self, request):
        lote_id_raw = str(request.data.get('lote_id') or '').strip()
        if not lote_id_raw:
            return Response(
                {'error': 'Se requiere el lote_id.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            lote_id = uuid.UUID(lote_id_raw)
        except ValueError:
            return Response(
                {'error': 'El lote_id enviado no es valido.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        preguntas = Pregunta.objects.filter(lote_id=lote_id)
        cantidad = preguntas.count()
        preguntas.delete()

        return Response(
            {
                'status': 'OK',
                'cantidad_eliminada': cantidad,
                'mensaje': f'Se eliminaron {cantidad} preguntas del lote {lote_id}.',
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'], url_path='criticas')
    def criticas(self, request):
        umbral_raw = str(request.query_params.get('umbral') or '60').strip()
        programa_id = str(request.query_params.get('programa_id') or '').strip()
        nivel = str(request.query_params.get('nivel') or '').strip()
        search = str(request.query_params.get('search') or '').strip()

        try:
            umbral = float(umbral_raw)
        except ValueError:
            return Response(
                {'detalle': 'El parametro umbral debe ser numerico.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filtro_base = Q(
            respuestaestudiante__intento__estado__in=['Finalizado', 'Pendiente Calificacion'],
        )

        if programa_id:
            filtro_base &= Q(
                respuestaestudiante__intento__estudiante__programa_id=programa_id,
            )

        preguntas = Pregunta.objects.all()
        if nivel:
            nivel_map = {
                'Facil': 'Facil',
                'Media': 'Medio',
                'Alta': 'Dificil',
                'Medio': 'Medio',
                'Dificil': 'Dificil',
            }
            nivel_backend = nivel_map.get(nivel, nivel)
            preguntas = preguntas.filter(nivel_dificultad=nivel_backend)

        if search:
            preguntas = preguntas.filter(enunciado__icontains=search)

        preguntas = (
            preguntas.select_related('modulo', 'categoria', 'competencia')
            .annotate(
                total_respondidas=Count('respuestaestudiante', filter=filtro_base),
                respuestas_incorrectas=Count(
                    'respuestaestudiante',
                    filter=(
                        filtro_base
                        & (
                            Q(respuestaestudiante__opcion_seleccionada__es_correcta=False)
                            | Q(respuestaestudiante__opcion_seleccionada__isnull=True)
                        )
                    ),
                ),
            )
            .annotate(
                tasa_error=Case(
                    When(total_respondidas=0, then=Value(0.0)),
                    default=ExpressionWrapper(
                        Cast(F('respuestas_incorrectas'), FloatField())
                        * Value(100.0)
                        / Cast(F('total_respondidas'), FloatField()),
                        output_field=FloatField(),
                    ),
                    output_field=FloatField(),
                ),
            )
            .filter(total_respondidas__gt=0, tasa_error__gte=umbral)
            .order_by('-tasa_error', '-total_respondidas', 'created_at')
        )

        preguntas_list = list(preguntas[:300])
        serializer = self.get_serializer(preguntas_list, many=True)

        data = []
        for index, item in enumerate(serializer.data):
            obj = preguntas_list[index]
            item['tasa_error'] = round(float(getattr(obj, 'tasa_error', 0.0) or 0.0), 2)
            item['total_respondidas'] = int(getattr(obj, 'total_respondidas', 0) or 0)
            item['respuestas_incorrectas'] = int(getattr(obj, 'respuestas_incorrectas', 0) or 0)
            data.append(item)

        return Response(
            {
                'umbral': umbral,
                'total': len(data),
                'results': data,
            },
            status=status.HTTP_200_OK,
        )


class ModuloAdminViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ModuloPrueba.objects.all().order_by('nombre')
    serializer_class = ModuloSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]


class CategoriaAdminViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = CategoriaAdminSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        queryset = Categoria.objects.select_related('modulo').all().order_by('nombre')
        modulo_id = self.request.query_params.get('modulo_id')
        if modulo_id:
            queryset = queryset.filter(modulo_id=modulo_id)
        return queryset

    def destroy(self, request, *args, **kwargs):
        categoria = self.get_object()

        if Pregunta.objects.filter(categoria=categoria).exists():
            return Response(
                {'detalle': 'No se puede eliminar la categoria porque tiene preguntas asociadas.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().destroy(request, *args, **kwargs)


class CompetenciaAdminViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = CompetenciaAdminSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        queryset = Competencia.objects.select_related('modulo').all().order_by('nombre')
        modulo_id = self.request.query_params.get('modulo_id')
        if modulo_id:
            queryset = queryset.filter(modulo_id=modulo_id)
        return queryset

    def destroy(self, request, *args, **kwargs):
        competencia = self.get_object()

        if Pregunta.objects.filter(competencia=competencia).exists():
            return Response(
                {'detalle': 'No se puede eliminar la competencia porque tiene preguntas asociadas.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().destroy(request, *args, **kwargs)


class PlantillaExamenAdminViewSet(viewsets.ModelViewSet):
    queryset = PlantillaExamen.objects.all()
    serializer_class = PlantillaExamenAdminSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        queryset = PlantillaExamen.objects.all()
        incluir_archivados = (
            str(self.request.query_params.get('incluir_archivados', 'false')).strip().lower()
            in ['1', 'true', 'yes']
        )

        if self.action == 'list' and not incluir_archivados:
            queryset = queryset.exclude(estado='Archivado')

        return queryset

    @staticmethod
    def _has_intentos(instance):
        if getattr(instance, 'intentos', None):
            return instance.intentos.exists()
        return IntentoExamen.objects.filter(plantilla_examen=instance).exists()

    @staticmethod
    def _resolver_nombre_estudiante(estudiante):
        nombres = str(getattr(estudiante, 'nombres', '') or '').strip()
        apellidos = str(getattr(estudiante, 'apellidos', '') or '').strip()
        full_name = f'{nombres} {apellidos}'.strip()
        if full_name:
            return full_name
        return str(getattr(estudiante, 'correo_institucional', '') or 'Estudiante')

    @staticmethod
    def _calcular_puntaje_intento(intento):
        respuestas = list(intento.respuestas.all())
        total_preguntas = len(respuestas)
        if total_preguntas == 0:
            return 0

        aciertos_multiples = sum(
            1
            for respuesta in respuestas
            if respuesta.opcion_seleccionada_id and respuesta.opcion_seleccionada.es_correcta
        )

        puntaje_ensayos = Decimal('0')
        for respuesta in respuestas:
            if respuesta.pregunta.limite_palabras is not None and respuesta.puntaje_calificado is not None:
                puntaje_ensayos += respuesta.puntaje_calificado

        total_aciertos = Decimal(aciertos_multiples) + puntaje_ensayos
        return round((float(total_aciertos) / total_preguntas) * 300)

    @staticmethod
    def _map_genero_label(genero):
        genero_map = {
            'M': 'Masculino',
            'F': 'Femenino',
            'O': 'Otro',
        }
        return genero_map.get(str(genero or '').strip().upper(), 'No especificado')

    @staticmethod
    def _get_filter_value(request, key):
        query_value = str(request.query_params.get(key) or '').strip()
        if query_value:
            return query_value

        data = getattr(request, 'data', None)
        if isinstance(data, dict):
            return str(data.get(key) or '').strip()
        return ''

    @staticmethod
    def _parse_bool(value):
        if isinstance(value, bool):
            return value
        return str(value or '').strip().lower() in ['1', 'true', 'yes', 'on']

    def _apply_result_filters(self, queryset, request):
        programa = self._get_filter_value(request, 'programa')
        genero = self._get_filter_value(request, 'genero').upper()
        semestre = self._get_filter_value(request, 'semestre')

        if programa:
            queryset = queryset.filter(estudiante__programa_id=programa)

        if genero in ['M', 'F', 'O']:
            queryset = queryset.filter(estudiante__genero=genero)

        if semestre and semestre.isdigit():
            queryset = queryset.filter(estudiante__semestre_actual=int(semestre))

        return queryset

    def _build_por_programa(self, resultados_data):
        programa_bucket = {}
        for row in resultados_data:
            programa_nombre = str(row.get('programa_nombre') or 'N/A').strip() or 'N/A'
            item = programa_bucket.setdefault(
                programa_nombre,
                {
                    'programa': programa_nombre,
                    'participantes': 0,
                    'acumulado': 0.0,
                },
            )
            item['participantes'] += 1
            item['acumulado'] += float(row.get('puntaje_global') or 0)

        por_programa = []
        for item in programa_bucket.values():
            participantes = int(item['participantes'] or 0)
            por_programa.append(
                {
                    'programa': item['programa'],
                    'promedio': round((item['acumulado'] / participantes), 2) if participantes else 0.0,
                    'participantes': participantes,
                }
            )

        por_programa.sort(
            key=lambda current: (
                -float(current.get('promedio') or 0),
                current.get('programa') or '',
            )
        )
        return por_programa

    def _build_resultados_data(self, intentos_queryset):
        data = []
        for intento in intentos_queryset:
            estudiante = intento.estudiante
            programa = getattr(estudiante, 'programa', None)
            puntaje_global = self._calcular_puntaje_intento(intento)
            data.append(
                {
                    'intento_id': str(intento.id),
                    'estudiante_nombre': self._resolver_nombre_estudiante(estudiante),
                    'programa_nombre': getattr(programa, 'nombre', None) or 'N/A',
                    'genero': str(getattr(estudiante, 'genero', '') or '').strip().upper() or None,
                    'genero_nombre': self._map_genero_label(getattr(estudiante, 'genero', None)),
                    'semestre': getattr(estudiante, 'semestre_actual', None),
                    'fecha_fin': intento.fecha_finalizacion.isoformat()
                    if intento.fecha_finalizacion
                    else None,
                    'puntaje_global': puntaje_global,
                }
            )

        data.sort(
            key=lambda item: (
                -float(item.get('puntaje_global', 0)),
                str(item.get('fecha_fin') or ''),
            )
        )

        for index, row in enumerate(data, start=1):
            row['posicion'] = index

        return data

    def _build_analiticas_data(self, intentos_queryset, resultados_data):
        intentos_ids = list(intentos_queryset.values_list('id', flat=True))
        if not intentos_ids:
            return {
                'promedio_global': 0.0,
                'por_dificultad': [],
                'por_competencia': [],
                'por_genero': [],
                'por_semestre': [],
                'rendimiento_preguntas': [],
            }

        acierto_expr = Case(
            When(opcion_seleccionada__es_correcta=True, then=Value(1.0)),
            When(
                Q(pregunta__limite_palabras__isnull=False) & Q(puntaje_calificado__gt=0),
                then=Value(1.0),
            ),
            default=Value(0.0),
            output_field=FloatField(),
        )

        respuestas = (
            RespuestaEstudiante.objects.filter(intento_id__in=intentos_ids)
            .select_related('pregunta')
            .annotate(acierto=acierto_expr)
        )

        promedio_global = (
            sum(float(item.get('puntaje_global') or 0) for item in resultados_data) / len(resultados_data)
            if resultados_data
            else 0.0
        )

        por_dificultad_rows = (
            respuestas.values('pregunta__nivel_dificultad')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('pregunta__nivel_dificultad')
        )
        por_dificultad = [
            {
                'dificultad': str(row.get('pregunta__nivel_dificultad') or 'Sin dato'),
                'promedio': round(float(row.get('promedio') or 0.0) * 300, 2),
                'participaciones': int(row.get('total') or 0),
            }
            for row in por_dificultad_rows
        ]

        por_competencia_rows = (
            respuestas.values('pregunta__competencia__nombre', 'pregunta__categoria__nombre')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('pregunta__competencia__nombre', 'pregunta__categoria__nombre')
        )
        por_competencia = [
            {
                'competencia': row.get('pregunta__competencia__nombre') or 'Sin competencia',
                'categoria': row.get('pregunta__categoria__nombre') or 'Sin categoria',
                'promedio': round(float(row.get('promedio') or 0.0) * 300, 2),
                'participaciones': int(row.get('total') or 0),
            }
            for row in por_competencia_rows
        ]

        genero_bucket = {}
        semestre_bucket = {}
        for row in resultados_data:
            genero_key = str(row.get('genero') or '').upper() or 'NA'
            genero_item = genero_bucket.setdefault(
                genero_key,
                {
                    'genero': genero_key if genero_key != 'NA' else None,
                    'genero_nombre': self._map_genero_label(genero_key),
                    'participantes': 0,
                    'acumulado': 0.0,
                },
            )
            genero_item['participantes'] += 1
            genero_item['acumulado'] += float(row.get('puntaje_global') or 0)

            semestre_key = row.get('semestre')
            semestre_label = semestre_key if semestre_key is not None else 'Sin dato'
            semestre_item = semestre_bucket.setdefault(
                semestre_label,
                {
                    'semestre': semestre_key,
                    'semestre_label': str(semestre_label),
                    'participantes': 0,
                    'acumulado': 0.0,
                },
            )
            semestre_item['participantes'] += 1
            semestre_item['acumulado'] += float(row.get('puntaje_global') or 0)

        por_genero = []
        for item in genero_bucket.values():
            participantes = int(item['participantes'] or 0)
            por_genero.append(
                {
                    'genero': item['genero'],
                    'genero_nombre': item['genero_nombre'],
                    'promedio': round((item['acumulado'] / participantes), 2) if participantes else 0.0,
                    'participantes': participantes,
                }
            )
        por_genero.sort(key=lambda item: item.get('genero_nombre') or '')

        por_semestre = []
        for item in semestre_bucket.values():
            participantes = int(item['participantes'] or 0)
            por_semestre.append(
                {
                    'semestre': item['semestre'],
                    'semestre_label': item['semestre_label'],
                    'promedio': round((item['acumulado'] / participantes), 2) if participantes else 0.0,
                    'participantes': participantes,
                }
            )
        por_semestre.sort(
            key=lambda item: (
                999 if item.get('semestre') is None else int(item.get('semestre')),
                item.get('semestre_label') or '',
            )
        )

        preguntas_rows = (
            respuestas.values('pregunta_id', 'pregunta__enunciado', 'pregunta__modulo__nombre')
            .annotate(
                total=Count('id'),
                promedio_acierto=Avg('acierto'),
            )
            .order_by('-total')
        )

        rendimiento_preguntas = []
        for row in preguntas_rows:
            promedio_acierto = float(row.get('promedio_acierto') or 0.0)
            acierto_pct = round(promedio_acierto * 100, 2)
            error_pct = round((1.0 - promedio_acierto) * 100, 2)
            rendimiento_preguntas.append(
                {
                    'pregunta_id': str(row.get('pregunta_id')),
                    'enunciado': row.get('pregunta__enunciado') or '',
                    'modulo': row.get('pregunta__modulo__nombre') or 'Sin modulo',
                    'participaciones': int(row.get('total') or 0),
                    'acierto_porcentaje': acierto_pct,
                    'error_porcentaje': error_pct,
                }
            )

        rendimiento_preguntas.sort(
            key=lambda item: (
                -float(item.get('error_porcentaje') or 0),
                -int(item.get('participaciones') or 0),
            )
        )

        return {
            'promedio_global': round(promedio_global, 2),
            'por_dificultad': por_dificultad,
            'por_competencia': por_competencia,
            'por_genero': por_genero,
            'por_semestre': por_semestre,
            'rendimiento_preguntas': rendimiento_preguntas[:10],
        }

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if self._has_intentos(instance):
            raise ValidationError(
                {'detail': 'No se puede editar un simulacro que ya tiene intentos de estudiantes.'}
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if self._has_intentos(instance):
            raise ValidationError(
                {'detail': 'No se puede eliminar un simulacro que ya tiene intentos de estudiantes.'}
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def archivar(self, request, pk=None):
        simulacro = self.get_object()
        simulacro.estado = 'Archivado'
        simulacro.save(update_fields=['estado', 'updated_at'])
        return Response({'detail': 'Simulacro archivado correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def resultados(self, request, pk=None):
        simulacro = self.get_object()
        intentos_qs = (
            IntentoExamen.objects.filter(plantilla_examen=simulacro, estado='Finalizado')
            .select_related('estudiante__programa')
            .prefetch_related('respuestas__opcion_seleccionada', 'respuestas__pregunta')
        )
        intentos = self._apply_result_filters(intentos_qs, request)
        data = self._build_resultados_data(intentos)

        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def analiticas_detalladas(self, request, pk=None):
        simulacro = self.get_object()
        intentos_qs = (
            IntentoExamen.objects.filter(plantilla_examen=simulacro, estado='Finalizado')
            .select_related('estudiante__programa')
            .prefetch_related('respuestas__opcion_seleccionada', 'respuestas__pregunta')
        )
        intentos = self._apply_result_filters(intentos_qs, request)
        resultados_data = self._build_resultados_data(intentos)
        analiticas = self._build_analiticas_data(intentos, resultados_data)

        return Response(
            {
                'simulacro_id': str(simulacro.id),
                'simulacro_titulo': simulacro.titulo,
                **analiticas,
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=True,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsAdminUser],
    )
    def exportar_excel_resultados(self, request, pk=None):
        simulacro = self.get_object()
        intentos_qs = (
            IntentoExamen.objects.filter(plantilla_examen=simulacro, estado='Finalizado')
            .select_related('estudiante__programa')
            .prefetch_related('respuestas__opcion_seleccionada', 'respuestas__pregunta')
        )
        intentos = self._apply_result_filters(intentos_qs, request)
        resultados_data = self._build_resultados_data(intentos)
        analiticas = self._build_analiticas_data(intentos, resultados_data)

        workbook = Workbook()
        sheet_resultados = workbook.active
        sheet_resultados.title = 'Resultados Generales'

        header_fill = PatternFill(fill_type='solid', fgColor='8F141B')
        header_font = Font(color='FFFFFF', bold=True)
        section_fill = PatternFill(fill_type='solid', fgColor='F3E6E8')
        section_font = Font(color='8F141B', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Puesto', 'Nombre', 'Programa', 'Semestre', 'Genero', 'Puntaje Global']
        sheet_resultados.append(headers)
        for column_index in range(1, len(headers) + 1):
            cell = sheet_resultados.cell(row=1, column=column_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for row in resultados_data:
            sheet_resultados.append(
                [
                    row.get('posicion'),
                    row.get('estudiante_nombre'),
                    row.get('programa_nombre'),
                    row.get('semestre') if row.get('semestre') is not None else 'Sin dato',
                    row.get('genero_nombre'),
                    row.get('puntaje_global'),
                ]
            )

        for column_cells in sheet_resultados.columns:
            max_length = 0
            for cell in column_cells:
                cell_value = '' if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            sheet_resultados.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 12),
                50,
            )

        sheet_analitica = workbook.create_sheet(title='Analitica Agrupada')

        def write_section(title, headers_section, rows, start_row):
            sheet_analitica.cell(row=start_row, column=1, value=title)
            title_cell = sheet_analitica.cell(row=start_row, column=1)
            title_cell.font = section_font
            title_cell.fill = section_fill

            header_row = start_row + 1
            for col_idx, header in enumerate(headers_section, start=1):
                cell = sheet_analitica.cell(row=header_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            current_row = header_row + 1
            for row_values in rows:
                for col_idx, value in enumerate(row_values, start=1):
                    sheet_analitica.cell(row=current_row, column=col_idx, value=value)
                current_row += 1

            return current_row + 1

        row_cursor = 1
        row_cursor = write_section(
            'Resumen General',
            ['Metrica', 'Valor'],
            [['Promedio Global (0-300)', analiticas.get('promedio_global', 0.0)]],
            row_cursor,
        )

        row_cursor = write_section(
            'Promedio por Dificultad',
            ['Dificultad', 'Promedio (0-300)', 'Participaciones'],
            [
                [item.get('dificultad'), item.get('promedio'), item.get('participaciones')]
                for item in analiticas.get('por_dificultad', [])
            ],
            row_cursor,
        )

        row_cursor = write_section(
            'Promedio por Competencia y Categoria',
            ['Competencia', 'Categoria', 'Promedio (0-300)', 'Participaciones'],
            [
                [
                    item.get('competencia'),
                    item.get('categoria'),
                    item.get('promedio'),
                    item.get('participaciones'),
                ]
                for item in analiticas.get('por_competencia', [])
            ],
            row_cursor,
        )

        row_cursor = write_section(
            'Desempeno por Genero',
            ['Genero', 'Promedio (0-300)', 'Participantes'],
            [
                [item.get('genero_nombre'), item.get('promedio'), item.get('participantes')]
                for item in analiticas.get('por_genero', [])
            ],
            row_cursor,
        )

        row_cursor = write_section(
            'Desempeno por Semestre',
            ['Semestre', 'Promedio (0-300)', 'Participantes'],
            [
                [item.get('semestre_label'), item.get('promedio'), item.get('participantes')]
                for item in analiticas.get('por_semestre', [])
            ],
            row_cursor,
        )

        write_section(
            'Rendimiento de Preguntas (Top 10 error)',
            ['Pregunta', 'Modulo', '% Acierto', '% Error', 'Participaciones'],
            [
                [
                    item.get('enunciado'),
                    item.get('modulo'),
                    item.get('acierto_porcentaje'),
                    item.get('error_porcentaje'),
                    item.get('participaciones'),
                ]
                for item in analiticas.get('rendimiento_preguntas', [])
            ],
            row_cursor,
        )

        for column_cells in sheet_analitica.columns:
            max_length = 0
            for cell in column_cells:
                cell_value = '' if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            sheet_analitica.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 14),
                60,
            )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="Reporte_Resultados_{simulacro.id}.xlsx"'
        )
        workbook.save(response)
        return response

    @action(detail=True, methods=['post'])
    def generar_reporte_excel_avanzado(self, request, pk=None):
        simulacro = self.get_object()

        # IDOR/Privilege Escalation check
        # Si no es Admin y no es el creador (suponiendo que haya un simulacro.creado_por), bloquear.
        # Validaremos si es Admin O Profesor. En el contexto de SaberPro, los roles suelen gestionarse así.
        from users.models import Usuario
        if request.user.rol not in [Usuario.ROL_ADMIN, Usuario.ROL_PROFESOR]:
            return Response({'detail': 'No tienes permisos para generar reportes avanzados.'}, status=status.HTTP_403_FORBIDDEN)

        # Payload Injection check: strict boolean validation
        payload_keys = ['incluir_general', 'incluir_programa', 'incluir_demografia', 'incluir_preguntas', 'incluir_competencias']
        for key in payload_keys:
            if key in request.data and not isinstance(request.data[key], bool):
                 return Response({'detail': f'Payload Invalido: {key} debe ser un booleano estricto.'}, status=status.HTTP_400_BAD_REQUEST)

        incluir_general = request.data.get('incluir_general', False)
        incluir_programa = request.data.get('incluir_programa', False)
        incluir_demografia = request.data.get('incluir_demografia', False)
        incluir_preguntas = request.data.get('incluir_preguntas', False)
        incluir_competencias = request.data.get('incluir_competencias', False)

        if not any([incluir_general, incluir_programa, incluir_demografia, incluir_preguntas, incluir_competencias]):
            return Response(
                {'detail': 'Debes seleccionar al menos una seccion para generar el reporte.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Prevencion de Memory Exhaustion: Quitar prefetch_related pesados y usar iterator()
        # Se usaran aggregate/annotate optimizados de SQL.
        intentos_qs = (
            IntentoExamen.objects.filter(plantilla_examen=simulacro, estado='Finalizado')
            .select_related('estudiante__programa')
        )
        intentos = self._apply_result_filters(intentos_qs, request)
        
        # Para resultados generales usamos .iterator() nativo 
        resultados_data_iterator = (
            self._calcular_puntaje_saber_pro(intento) for intento in intentos.iterator(chunk_size=1000)
        )
        # Como build_analiticas y otros asumen tener resultados_data en memoria local (list),
        # por ahora para resolver DoS extremo reusamos la lista pero limitando el prefetch:
        resultados_data = self._build_resultados_data(intentos.iterator(chunk_size=1000))
        analiticas = self._build_analiticas_data(intentos, resultados_data)
        por_programa = self._build_por_programa(resultados_data)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        header_fill = PatternFill(fill_type='solid', fgColor='8F141B')
        header_font = Font(color='FFFFFF', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        def style_header(sheet, headers):
            sheet.append(headers)
            for column_index in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=column_index)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

        def autosize(sheet, min_width=12, max_width=70):
            for column_cells in sheet.columns:
                max_length = 0
                for cell in column_cells:
                    cell_value = '' if cell.value is None else str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(
                    max(max_length + 2, min_width),
                    max_width,
                )

        if incluir_general:
            sheet = workbook.create_sheet(title='Resultados Generales')
            headers = ['Puesto', 'Nombre', 'Programa', 'Semestre', 'Genero', 'Puntaje Global']
            style_header(sheet, headers)

            for row in resultados_data:
                sheet.append(
                    [
                        row.get('posicion'),
                        row.get('estudiante_nombre'),
                        row.get('programa_nombre'),
                        row.get('semestre') if row.get('semestre') is not None else 'Sin dato',
                        row.get('genero_nombre'),
                        row.get('puntaje_global'),
                    ]
                )
            autosize(sheet)

        if incluir_programa:
            sheet = workbook.create_sheet(title='Por Programa')
            headers = ['Programa', 'Promedio (0-300)', 'Participantes']
            style_header(sheet, headers)
            for item in por_programa:
                sheet.append([item.get('programa'), item.get('promedio'), item.get('participantes')])
            autosize(sheet)

        if incluir_demografia:
            sheet = workbook.create_sheet(title='Demografia')
            headers = ['Tipo', 'Segmento', 'Promedio (0-300)', 'Participantes']
            style_header(sheet, headers)

            for item in analiticas.get('por_genero', []):
                sheet.append(
                    [
                        'Genero',
                        item.get('genero_nombre'),
                        item.get('promedio'),
                        item.get('participantes'),
                    ]
                )

            for item in analiticas.get('por_semestre', []):
                sheet.append(
                    [
                        'Semestre',
                        item.get('semestre_label'),
                        item.get('promedio'),
                        item.get('participantes'),
                    ]
                )
            autosize(sheet)

        if incluir_competencias:
            sheet = workbook.create_sheet(title='Competencias Categorias')
            headers = ['Competencia', 'Categoria', 'Promedio (0-300)', 'Participaciones']
            style_header(sheet, headers)
            for item in analiticas.get('por_competencia', []):
                sheet.append(
                    [
                        item.get('competencia'),
                        item.get('categoria'),
                        item.get('promedio'),
                        item.get('participaciones'),
                    ]
                )
            autosize(sheet)

        if incluir_preguntas:
            sheet = workbook.create_sheet(title='Analisis de Preguntas')
            headers = [
                'Pregunta',
                'Modulo',
                'Total Respuestas',
                'Aciertos',
                'Errores',
                'Tasa Acierto (%)',
                'Tasa Error (%)',
                'Clasificacion',
            ]
            style_header(sheet, headers)

            intentos_ids = list(intentos.values_list('id', flat=True))
            acierto_expr = Case(
                When(opcion_seleccionada__es_correcta=True, then=Value(1.0)),
                When(
                    Q(pregunta__limite_palabras__isnull=False) & Q(puntaje_calificado__gt=0),
                    then=Value(1.0),
                ),
                default=Value(0.0),
                output_field=FloatField(),
            )

            preguntas_rows = (
                RespuestaEstudiante.objects.filter(intento_id__in=intentos_ids)
                .annotate(acierto=acierto_expr)
                .values('pregunta__enunciado', 'pregunta__modulo__nombre')
                .annotate(total_respuestas=Count('id'), tasa_acierto=Avg('acierto'))
                .order_by('-total_respuestas')
            )

            ranking = []
            for row in preguntas_rows:
                total_respuestas = int(row.get('total_respuestas') or 0)
                tasa_acierto_ratio = float(row.get('tasa_acierto') or 0.0)
                aciertos = round(total_respuestas * tasa_acierto_ratio)
                errores = max(total_respuestas - aciertos, 0)
                tasa_acierto_pct = round(tasa_acierto_ratio * 100, 2)
                tasa_error_pct = round((1.0 - tasa_acierto_ratio) * 100, 2)
                ranking.append(
                    {
                        'enunciado': row.get('pregunta__enunciado') or '',
                        'modulo': row.get('pregunta__modulo__nombre') or 'Sin modulo',
                        'total_respuestas': total_respuestas,
                        'aciertos': aciertos,
                        'errores': errores,
                        'tasa_acierto_pct': tasa_acierto_pct,
                        'tasa_error_pct': tasa_error_pct,
                    }
                )

            faciles = sorted(
                ranking,
                key=lambda item: (
                    -float(item.get('tasa_acierto_pct') or 0),
                    -int(item.get('total_respuestas') or 0),
                ),
            )[:10]
            dificiles = sorted(
                ranking,
                key=lambda item: (
                    -float(item.get('tasa_error_pct') or 0),
                    -int(item.get('total_respuestas') or 0),
                ),
            )[:10]

            for item in dificiles:
                sheet.append(
                    [
                        item.get('enunciado'),
                        item.get('modulo'),
                        item.get('total_respuestas'),
                        item.get('aciertos'),
                        item.get('errores'),
                        item.get('tasa_acierto_pct'),
                        item.get('tasa_error_pct'),
                        'Mas dificil',
                    ]
                )

            for item in faciles:
                sheet.append(
                    [
                        item.get('enunciado'),
                        item.get('modulo'),
                        item.get('total_respuestas'),
                        item.get('aciertos'),
                        item.get('errores'),
                        item.get('tasa_acierto_pct'),
                        item.get('tasa_error_pct'),
                        'Mas facil',
                    ]
                )

            autosize(sheet, max_width=85)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="Reporte_Avanzado_Simulacro_{simulacro.id}.xlsx"'
        )
        workbook.save(response)
        return response

    @action(detail=True, methods=['get'])
    def descargar_muestra(self, request, pk=None):
        simulacro = self.get_object()

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = (
            f'attachment; filename="Muestra_Simulacro_{simulacro.id}.csv"'
        )
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';')
        writer.writerow(
            [
                'Modulo',
                'Categoria',
                'Dificultad',
                'Enunciado',
                'Opcion Correcta',
                'Distractor 1',
                'Distractor 2',
                'Distractor 3',
            ]
        )

        preguntas_agregadas = set()

        for regla in simulacro.reglas.select_related('modulo', 'categoria').all():
            preguntas_qs = Pregunta.objects.filter(
                modulo=regla.modulo,
                estado='Publicada',
            )

            if regla.nivel_dificultad and regla.nivel_dificultad != 'Balanceada':
                preguntas_qs = preguntas_qs.filter(nivel_dificultad=regla.nivel_dificultad)

            if regla.categoria_id:
                preguntas_qs = preguntas_qs.filter(categoria=regla.categoria)

            if preguntas_agregadas:
                preguntas_qs = preguntas_qs.exclude(id__in=preguntas_agregadas)

            preguntas = list(
                preguntas_qs.select_related('categoria')
                .prefetch_related('opciones')
                .distinct()
                .order_by('?')[: regla.cantidad_preguntas]
            )

            if len(preguntas) < regla.cantidad_preguntas:
                raise ValidationError(
                    {
                        'detail': (
                            f'No hay suficientes preguntas publicadas para la regla {regla.id}. '
                            'Ajusta las reglas o publica mas preguntas antes de descargar la muestra.'
                        )
                    }
                )

            for pregunta in preguntas:
                preguntas_agregadas.add(pregunta.id)
                opciones = list(pregunta.opciones.all())
                opcion_correcta = next(
                    (
                        str(opcion.texto or '').strip()
                        for opcion in opciones
                        if opcion.es_correcta and str(opcion.texto or '').strip()
                    ),
                    '',
                )
                distractores = [
                    str(opcion.texto or '').strip()
                    for opcion in opciones
                    if not opcion.es_correcta and str(opcion.texto or '').strip()
                ]
                while len(distractores) < 3:
                    distractores.append('')

                writer.writerow(
                    [
                        str(regla.modulo.nombre or '').strip(),
                        str(getattr(pregunta.categoria, 'nombre', '') or '').strip(),
                        str(pregunta.nivel_dificultad or '').strip(),
                        str(pregunta.enunciado or '').strip(),
                        opcion_correcta,
                        distractores[0],
                        distractores[1],
                        distractores[2],
                    ]
                )

        return response


class EstudianteExamenViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PlantillaExamenEstudianteSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        programa_id = getattr(self.request.user, 'programa_id', None)
        now = timezone.now()
        queryset = PlantillaExamen.objects.prefetch_related('reglas').filter(
            estado='Activo',
            fecha_inicio__lte=now,
            fecha_fin__gte=now,
        )

        if programa_id:
            return queryset.filter(
                Q(programas_destino=programa_id) | Q(programas_destino__isnull=True)
            ).distinct()

        return queryset.filter(programas_destino__isnull=True).distinct()

    @action(detail=True, methods=['post'])
    def iniciar_intento(self, request, pk=None):
        plantilla = self.get_object()

        try:
            with transaction.atomic():
                intento, created = IntentoExamen.objects.get_or_create(
                    estudiante=request.user,
                    plantilla_examen=plantilla,
                    defaults={'estado': 'En Progreso'}
                )

                if not created:
                    return Response({'detalle': 'Intento ya iniciado'}, status=status.HTTP_400_BAD_REQUEST)

                modulos_creados = set()
                preguntas_agregadas = set()

                for regla in plantilla.reglas.select_related('modulo', 'categoria').all():
                    preguntas_qs = Pregunta.objects.filter(
                        modulo=regla.modulo,
                        estado='Publicada',
                    )

                    if regla.nivel_dificultad and regla.nivel_dificultad != 'Balanceada':
                        preguntas_qs = preguntas_qs.filter(nivel_dificultad=regla.nivel_dificultad)

                    if regla.categoria_id:
                        preguntas_qs = preguntas_qs.filter(categoria=regla.categoria)

                    if preguntas_agregadas:
                        preguntas_qs = preguntas_qs.exclude(id__in=preguntas_agregadas)

                    preguntas = list(
                        preguntas_qs.distinct().order_by('?')[: regla.cantidad_preguntas]
                    )

                    if len(preguntas) < regla.cantidad_preguntas:
                        raise ValueError(
                            f'No hay suficientes preguntas publicadas para la regla {regla.id}.'
                        )

                    if regla.modulo_id not in modulos_creados:
                        IntentoModulo.objects.create(
                            intento=intento,
                            modulo=regla.modulo,
                            estado='Pendiente',
                        )
                        modulos_creados.add(regla.modulo_id)

                    for pregunta in preguntas:
                        if pregunta.id in preguntas_agregadas:
                            continue
                        RespuestaEstudiante.objects.create(
                            intento=intento,
                            pregunta=pregunta,
                        )
                        preguntas_agregadas.add(pregunta.id)

        except ValueError as exc:
            return Response({'detalle': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'intento_id': str(intento.id)}, status=status.HTTP_201_CREATED)


class IntentoExamenViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = IntentoExamenSerializer

    def get_queryset(self):
        return (
            IntentoExamen.objects.filter(estudiante=self.request.user)
            .select_related('plantilla_examen')
            .prefetch_related('respuestas__opcion_seleccionada', 'respuestas__pregunta')
        )

    @staticmethod
    def _calcular_puntaje_saber_pro(intento):
        respuestas = list(intento.respuestas.all())
        total_preguntas = len(respuestas)
        if total_preguntas == 0:
            return 0

        aciertos_multiples = 0
        puntaje_ensayos = Decimal('0')

        for respuesta in respuestas:
            if respuesta.opcion_seleccionada_id and respuesta.opcion_seleccionada.es_correcta:
                aciertos_multiples += 1

            if respuesta.pregunta.limite_palabras is not None and respuesta.puntaje_calificado is not None:
                puntaje_ensayos += respuesta.puntaje_calificado

        total_aciertos = Decimal(aciertos_multiples) + puntaje_ensayos
        return round((float(total_aciertos) / total_preguntas) * 300)

    @action(detail=True, methods=['get'])
    def cargar_respuestas(self, request, pk=None):
        intento = self.get_object()
        modulo_id = request.query_params.get('modulo')

        respuestas = (
            RespuestaEstudiante.objects.filter(intento=intento)
            .select_related('pregunta')
            .order_by('id')
        )
        if modulo_id:
            respuestas = respuestas.filter(pregunta__modulo_id=modulo_id)

        serializer = RespuestaEstudianteDetalleSerializer(respuestas, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='preguntas')
    def preguntas(self, request, pk=None):
        intento = self.get_object()
        respuestas = (
            RespuestaEstudiante.objects.filter(intento=intento)
            .select_related('pregunta')
            .order_by('id')
        )
        serializer = RespuestaEstudianteDetalleSerializer(respuestas, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def finalizar(self, request, pk=None):
        intento = self.get_object()

        if intento.estado == 'En Progreso':
            ensayos_pendientes = intento.respuestas.filter(
                pregunta__limite_palabras__isnull=False,
                puntaje_calificado__isnull=True,
            ).exists()

            intento.estado = 'Pendiente Calificacion' if ensayos_pendientes else 'Finalizado'
            intento.fecha_finalizacion = timezone.now()
            intento.save()

        return Response({'estado': intento.estado})

    @action(detail=True, methods=['post'])
    def generar_plan_estudio(self, request, pk=None):
        intento = self.get_object()

        if intento.plan_estudio_ia:
            return Response({'plan': intento.plan_estudio_ia}, status=status.HTTP_200_OK)

        respuestas = intento.respuestas.select_related(
            'pregunta__modulo',
            'pregunta__competencia',
            'pregunta__categoria',
            'opcion_seleccionada',
        )

        respuestas_malas = []
        for respuesta in respuestas:
            pregunta = respuesta.pregunta

            if pregunta.limite_palabras is None:
                # Opcion multiple: sin seleccionar o seleccion incorrecta cuenta como falla.
                if not respuesta.opcion_seleccionada_id or not respuesta.opcion_seleccionada.es_correcta:
                    respuestas_malas.append(respuesta)
                continue

            # Ensayo: solo se considera falla si ya fue calificado en 0 o menos.
            if respuesta.puntaje_calificado is not None and float(respuesta.puntaje_calificado) <= 0:
                respuestas_malas.append(respuesta)

        if len(respuestas_malas) == 0:
            mensaje = (
                '## Excelente resultado\n\n'
                'Felicidades, no se identificaron respuestas incorrectas en este intento. '
                'Sigue practicando para mantener ese nivel.'
            )
            intento.plan_estudio_ia = mensaje
            intento.save(update_fields=['plan_estudio_ia'])
            return Response({'plan': mensaje}, status=status.HTTP_200_OK)

        debilidades = {}
        for respuesta in respuestas_malas:
            modulo_nombre = respuesta.pregunta.modulo.nombre
            competencia_nombre = (
                respuesta.pregunta.competencia.nombre
                if respuesta.pregunta.competencia_id
                else 'General'
            )
            debilidades.setdefault(modulo_nombre, set()).add(competencia_nombre)

        resumen_fallos = '\n'.join(
            [
                f"- Modulo {modulo}: Fallo en competencias como {', '.join(sorted(competencias))}"
                for modulo, competencias in debilidades.items()
            ]
        )

        puntaje_global = self._calcular_puntaje_saber_pro(intento)

        prompt = f"""
      Eres un tutor experto en pruebas ICFES Saber Pro. Un estudiante ha fallado preguntas en tu área.

      Puntaje obtenido: {puntaje_global}
      Áreas de fallo exactas: {resumen_fallos}

      Tu objetivo no es darle consejos generales ni motivación vacía. Tu objetivo es ENSEÑARLE a resolver sus errores AHORA MISMO.
      Crea una micro-lección de tutoría en formato Markdown estructurada estrictamente de la siguiente manera:

      ### 💡 ¿Por qué es importante esto?
      (Explica en un párrafo corto y directo, sin rodeos, qué significa esta competencia en la vida real o en la prueba).

      ### 🔍 Concepto Clave (Repaso Rápido)
      (Explica el concepto matemático, de lectura o lógica que el estudiante necesita saber para no volver a fallar. Usa ejemplos cortos. Si falló en gráficas, explícale cómo leer los ejes X y Y).

      ### 🏋️‍♂️ Gimnasio Mental: Ejercicios Prácticos
      (Crea 3 ejercicios de práctica RÁPIDOS y directos basados en las competencias donde falló. Plantea el problema y haz la pregunta).
      * **Ejercicio 1:** [Planteamiento del problema de nivel básico]
      * **Ejercicio 2:** [Planteamiento del problema de nivel medio]
      * **Ejercicio 3:** [Planteamiento del problema de nivel alto]

      ### ✅ Respuestas Explicadas
      (Proporciona la solución paso a paso de los 3 ejercicios anteriores para que el estudiante pueda autoevaluarse inmediatamente).

      No incluyas cronogramas de días, no lo mandes a buscar en otras páginas, no uses frases cliché extensas. Sé un profesor al grano, práctico y enfocado en la resolución de problemas.
      """

        try:
            genai.configure(api_key=settings.GEMINI_API_KEY)
            modelo_nombre, _ = GenerarOpcionesIAView._resolver_modelo_gemini()
            model = genai.GenerativeModel(modelo_nombre)
            response = model.generate_content(prompt)
            plan_generado = str(getattr(response, 'text', '') or '').strip()

            if not plan_generado:
                raise ValueError('La IA no devolvio contenido para el plan.')

            intento.plan_estudio_ia = plan_generado
            intento.save(update_fields=['plan_estudio_ia'])

            return Response({'plan': plan_generado}, status=status.HTTP_200_OK)

        except Exception as exc:
            return Response(
                {
                    'detalle': 'No fue posible generar el plan de estudio en este momento.',
                    'error': str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=['get'])
    def resumen_resultados(self, request, pk=None):
        try:
            intento = (
                IntentoExamen.objects.select_related('plantilla_examen')
                .prefetch_related('respuestas__opcion_seleccionada', 'respuestas__pregunta')
                .get(pk=pk, estudiante=request.user)
            )
        except IntentoExamen.DoesNotExist:
            return Response({'detalle': 'Intento no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        plantilla = intento.plantilla_examen
        now = timezone.now()
        if not plantilla.mostrar_resultados_inmediatos and now < plantilla.fecha_fin:
            return Response(
                {
                    'detalle': 'Los resultados estaran disponibles cuando cierre el examen.',
                    'fecha_disponible': plantilla.fecha_fin.isoformat(),
                    'simulacro': plantilla.titulo,
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        respuestas = list(intento.respuestas.all())
        total_preguntas = len(respuestas)
        ensayos_pendientes = intento.respuestas.filter(
            pregunta__limite_palabras__isnull=False,
            puntaje_calificado__isnull=True,
        ).exists()

        if total_preguntas == 0:
            return Response(
                {
                    'intento_id': str(intento.id),
                    'estado_calificacion': 'Parcial' if ensayos_pendientes else 'Definitiva',
                    'puntaje_saber_pro': 0,
                    'total_preguntas': 0,
                    'aciertos_brutos': 0,
                    'detalle_respuestas': [],
                    'puntajes_por_modulo': [],
                    'plan_estudio_ia': intento.plan_estudio_ia,
                },
                status=status.HTTP_200_OK,
            )

        aciertos_multiples = 0
        puntaje_ensayos = Decimal('0')

        for respuesta in respuestas:
            if respuesta.opcion_seleccionada_id and respuesta.opcion_seleccionada.es_correcta:
                aciertos_multiples += 1

            if respuesta.pregunta.limite_palabras is not None and respuesta.puntaje_calificado is not None:
                puntaje_ensayos += respuesta.puntaje_calificado

        total_aciertos = Decimal(aciertos_multiples) + puntaje_ensayos
        puntaje_normalizado = round((float(total_aciertos) / total_preguntas) * 300)

        modulos_data = []
        modulos_stats = intento.respuestas.values('pregunta__modulo__nombre').annotate(
            total=Count('id'),
            aciertos=Count('id', filter=Q(opcion_seleccionada__es_correcta=True)),
        )

        for mod in modulos_stats:
            nombre = mod['pregunta__modulo__nombre'] or 'General'
            total_mod = mod['total']
            aciertos_mod = mod['aciertos']
            puntaje_mod = round((aciertos_mod / total_mod) * 300) if total_mod > 0 else 0
            percentil = round((puntaje_mod / 300) * 100)
            modulos_data.append(
                {
                    'modulo': nombre,
                    'puntaje': puntaje_mod,
                    'percentil': percentil,
                }
            )

        return Response(
            {
                'intento_id': str(intento.id),
                'estado_calificacion': 'Parcial' if ensayos_pendientes else 'Definitiva',
                'puntaje_saber_pro': puntaje_normalizado,
                'total_preguntas': total_preguntas,
                'aciertos_brutos': float(total_aciertos),
                'detalle_respuestas': RevisionRespuestaSerializer(respuestas, many=True).data,
                'puntajes_por_modulo': modulos_data,
                'plan_estudio_ia': intento.plan_estudio_ia,
            },
            status=status.HTTP_200_OK,
        )


class RespuestaEstudianteViewSet(mixins.UpdateModelMixin, viewsets.GenericViewSet):
    queryset = RespuestaEstudiante.objects.all()
    serializer_class = RespuestaEstudianteUpdateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return RespuestaEstudiante.objects.filter(intento__estudiante=self.request.user)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            try:
                # Bloqueo DB (select_for_update) contra Race Conditions.
                # Refuerzo IDOR solicitando coincidencia exacta de usuario logueado en DB.
                respuesta = RespuestaEstudiante.objects.select_related(
                    'intento__plantilla_examen'
                ).select_for_update().get(
                    pk=kwargs.get('pk'),
                    intento__estudiante=request.user
                )
            except RespuestaEstudiante.DoesNotExist:
                return Response(
                    {'detail': 'Respuesta no encontrada o acceso denegado.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            if respuesta.intento.estado != 'En Progreso':
                return Response(
                    {'detail': 'El intento ya fue finalizado y no permite cambios en las respuestas.'},
                    status=status.HTTP_409_CONFLICT,
                )

            if respuesta.intento.plantilla_examen.fecha_fin < timezone.now():
                return Response(
                    {'detail': 'El tiempo del examen ha expirado. No se aceptan más respuestas.'},
                    status=status.HTTP_409_CONFLICT,
                )

            serializer = self.get_serializer(respuesta, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)


class EvaluacionEnsayoViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = EnsayoPendienteSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return RespuestaEstudiante.objects.filter(
            pregunta__limite_palabras__isnull=False,
            texto_respuesta_abierta__isnull=False,
            puntaje_calificado__isnull=True,
        ).select_related('pregunta', 'evaluador')

    @action(detail=True, methods=['post'])
    def asignar_ensayo(self, request, pk=None):
        with transaction.atomic():
            try:
                ensayo = self.get_queryset().select_for_update().get(pk=pk)
            except RespuestaEstudiante.DoesNotExist:
                return Response({'detalle': 'Ensayo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

            if ensayo.evaluador and ensayo.evaluador != request.user:
                return Response(
                    {'detalle': 'Este ensayo ya esta siendo evaluado por otro profesor'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            ensayo.evaluador = request.user
            ensayo.save(update_fields=['evaluador'])

        return Response(
            {'mensaje': 'Ensayo asignado exitosamente'},
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'])
    def calificar(self, request, pk=None):
        serializer = CalificarEnsayoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ensayo = self.get_object()

        if ensayo.evaluador != request.user:
            return Response(
                {'detalle': 'No tienes permiso para calificar este ensayo'},
                status=status.HTTP_403_FORBIDDEN,
            )

        ensayo.puntaje_calificado = serializer.validated_data['puntaje']
        ensayo.save(update_fields=['puntaje_calificado'])

        return Response(
            {'mensaje': 'Calificacion guardada'},
            status=status.HTTP_200_OK,
        )


class SimulacrosDashboardStatsView(APIView):
    permission_classes = [IsAdminUser]

    @staticmethod
    def _build_objetivo_por_simulacro():
        rows = (
            Usuario.objects.filter(
                rol=Usuario.ROL_ESTUDIANTE,
                is_active=True,
                programa__isnull=False,
                programa__examenes_asignados__estado='Activo',
            )
            .values('programa__examenes_asignados')
            .annotate(total=Count('id', distinct=True))
        )
        return {
            str(row['programa__examenes_asignados']): int(row['total'] or 0)
            for row in rows
            if row.get('programa__examenes_asignados')
        }

    @staticmethod
    def _build_completados_por_simulacro():
        rows = (
            IntentoExamen.objects.filter(
                plantilla_examen__estado='Activo',
                estado='Finalizado',
                estudiante__rol=Usuario.ROL_ESTUDIANTE,
                estudiante__is_active=True,
            )
            .values('plantilla_examen')
            .annotate(total=Count('estudiante_id', distinct=True))
        )
        return {
            str(row['plantilla_examen']): int(row['total'] or 0)
            for row in rows
            if row.get('plantilla_examen')
        }

    @staticmethod
    def _build_promedio_puntaje_por_simulacro():
        acierto_expr = Case(
            When(opcion_seleccionada__es_correcta=True, then=Value(1.0)),
            When(
                Q(pregunta__limite_palabras__isnull=False) & Q(puntaje_calificado__gt=0),
                then=Value(1.0),
            ),
            default=Value(0.0),
            output_field=FloatField(),
        )

        rows = (
            RespuestaEstudiante.objects.filter(
                intento__plantilla_examen__estado='Activo',
                intento__estado='Finalizado',
                intento__estudiante__rol=Usuario.ROL_ESTUDIANTE,
                intento__estudiante__is_active=True,
            )
            .annotate(acierto=acierto_expr)
            .values('intento__plantilla_examen')
            .annotate(promedio=Avg('acierto'))
        )

        result = {}
        for row in rows:
            simulacro_id = row.get('intento__plantilla_examen')
            promedio = row.get('promedio')
            if not simulacro_id or promedio is None:
                continue
            result[str(simulacro_id)] = round(float(promedio) * 300, 2)
        return result

    @staticmethod
    def _build_promedio_tiempo_por_simulacro():
        rows = (
            IntentoExamen.objects.filter(
                plantilla_examen__estado='Activo',
                estado='Finalizado',
                fecha_finalizacion__isnull=False,
                estudiante__rol=Usuario.ROL_ESTUDIANTE,
                estudiante__is_active=True,
            )
            .annotate(
                duracion=ExpressionWrapper(
                    F('fecha_finalizacion') - F('fecha_inicio'),
                    output_field=DurationField(),
                )
            )
            .values('plantilla_examen')
            .annotate(promedio=Avg('duracion'))
        )

        result = {}
        for row in rows:
            simulacro_id = row.get('plantilla_examen')
            promedio = row.get('promedio')
            if not simulacro_id or promedio is None:
                continue
            result[str(simulacro_id)] = round(float(promedio.total_seconds()) / 60, 2)
        return result

    def get(self, request):
        simulacros_activos = PlantillaExamen.objects.filter(estado='Activo').order_by('fecha_inicio', 'titulo')
        total_simulacros = PlantillaExamen.objects.count()
        total_activos = simulacros_activos.count()

        objetivo_por_simulacro = self._build_objetivo_por_simulacro()
        completados_por_simulacro = self._build_completados_por_simulacro()
        promedio_puntaje_por_simulacro = self._build_promedio_puntaje_por_simulacro()
        promedio_tiempo_por_simulacro = self._build_promedio_tiempo_por_simulacro()

        data_simulacros_activos = []
        for simulacro in simulacros_activos:
            simulacro_id = str(simulacro.id)
            poblacion_objetivo = int(objetivo_por_simulacro.get(simulacro_id, 0) or 0)
            completados = int(completados_por_simulacro.get(simulacro_id, 0) or 0)
            pendientes = max(poblacion_objetivo - completados, 0)
            porcentaje_completado = (
                round((min(completados, poblacion_objetivo) / poblacion_objetivo) * 100, 2)
                if poblacion_objetivo > 0
                else 0.0
            )

            data_simulacros_activos.append(
                {
                    'id': simulacro_id,
                    'nombre': simulacro.titulo,
                    'poblacion_objetivo': poblacion_objetivo,
                    'completados': completados,
                    'pendientes': pendientes,
                    'porcentaje_completado': porcentaje_completado,
                    'promedio_puntaje': promedio_puntaje_por_simulacro.get(simulacro_id),
                    'promedio_tiempo_minutos': promedio_tiempo_por_simulacro.get(simulacro_id),
                }
            )

        return Response(
            {
                'globales': {
                    'total': total_simulacros,
                    'activos': total_activos,
                },
                'simulacros_activos': data_simulacros_activos,
            },
            status=status.HTTP_200_OK,
        )


class AnaliticasAdminViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]

    @action(detail=False, methods=['get'])
    def kpis_globales(self, request):
        simulacro_id = request.query_params.get('simulacro_id')

        intentos = IntentoExamen.objects.filter(estado__in=['Finalizado', 'Pendiente Calificacion'])
        if simulacro_id:
            intentos = intentos.filter(plantilla_examen_id=simulacro_id)

        total_intentos = intentos.count()

        participacion_programas = (
            intentos.values('estudiante__programa__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')
        )

        filtro_respuestas = Q(
            respuestaestudiante__intento__estado__in=['Finalizado', 'Pendiente Calificacion'],
        )
        if simulacro_id:
            filtro_respuestas &= Q(
                respuestaestudiante__intento__plantilla_examen_id=simulacro_id,
            )

        preguntas_anotadas = Pregunta.objects.annotate(
            total_respuestas=Count('respuestaestudiante', filter=filtro_respuestas),
            respuestas_incorrectas=Count(
                'respuestaestudiante',
                filter=(
                    filtro_respuestas
                    & (
                        Q(respuestaestudiante__opcion_seleccionada__es_correcta=False)
                        | Q(respuestaestudiante__opcion_seleccionada__isnull=True)
                    )
                ),
            ),
        ).filter(total_respuestas__gt=0)

        preguntas_criticas_qs = preguntas_anotadas.annotate(
            tasa_error=ExpressionWrapper(
                Cast(F('respuestas_incorrectas'), FloatField())
                * Value(100.0)
                / Cast(F('total_respuestas'), FloatField()),
                output_field=FloatField(),
            )
        ).filter(tasa_error__gte=60.0)

        total_preguntas_criticas = preguntas_criticas_qs.count()
        tasa_media_error_criticas = (
            preguntas_criticas_qs.aggregate(promedio=Avg('tasa_error')).get('promedio') or 0.0
        )

        top_preguntas_criticas_raw = (
            preguntas_criticas_qs.values(
                'enunciado',
                'total_respuestas',
                'respuestas_incorrectas',
                'tasa_error',
            )
            .order_by('-tasa_error', '-respuestas_incorrectas', 'created_at')[:5]
        )

        top_preguntas_criticas = [
            {
                'pregunta__enunciado': item['enunciado'],
                'total_veces': int(item['total_respuestas'] or 0),
                'veces_incorrecta': int(item['respuestas_incorrectas'] or 0),
                'tasa_error': round(float(item['tasa_error'] or 0.0), 2),
            }
            for item in top_preguntas_criticas_raw
        ]

        return Response(
            {
                'total_evaluaciones_finalizadas': total_intentos,
                'total_preguntas_criticas': total_preguntas_criticas,
                'tasa_media_error_criticas': round(float(tasa_media_error_criticas), 2),
                'participacion_por_programa': list(participacion_programas),
                'top_preguntas_criticas': top_preguntas_criticas,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'], url_path='cobertura-programa')
    def cobertura_programa(self, request):
        cobertura = (
            ProgramaAcademico.objects.annotate(
                total_estudiantes=Count(
                    'estudiantes',
                    filter=Q(estudiantes__rol=Usuario.ROL_ESTUDIANTE, estudiantes__is_active=True),
                    distinct=True,
                ),
                estudiantes_con_intento=Count(
                    'estudiantes',
                    filter=Q(
                        estudiantes__rol=Usuario.ROL_ESTUDIANTE,
                        estudiantes__is_active=True,
                        estudiantes__intentos__estado='Finalizado',
                    ),
                    distinct=True,
                ),
            )
            .order_by('nombre')
        )

        data = []
        for programa in cobertura:
            total = int(getattr(programa, 'total_estudiantes', 0) or 0)
            con_intento = int(getattr(programa, 'estudiantes_con_intento', 0) or 0)
            porcentaje = round((con_intento / total) * 100, 2) if total > 0 else 0.0
            data.append(
                {
                    'programa_id': programa.id,
                    'programa_nombre': programa.nombre,
                    'total_estudiantes_activos': total,
                    'estudiantes_con_intento_finalizado': con_intento,
                    'cobertura_porcentaje': porcentaje,
                }
            )

        return Response({'results': data}, status=status.HTTP_200_OK)


class ReportesViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]

    @staticmethod
    def _base_respuestas(request):
        programa_id = str(request.query_params.get('programa_id') or '').strip()
        simulacro_id = str(request.query_params.get('simulacro_id') or '').strip()

        filtros = Q(intento__estado__in=['Finalizado', 'Pendiente Calificacion'])
        if programa_id:
            filtros &= Q(intento__estudiante__programa_id=programa_id)
        if simulacro_id:
            filtros &= Q(intento__plantilla_examen_id=simulacro_id)

        acierto_expr = Case(
            When(opcion_seleccionada__es_correcta=True, then=Value(1.0)),
            default=Value(0.0),
            output_field=FloatField(),
        )

        return (
            RespuestaEstudiante.objects.filter(filtros)
            .select_related('pregunta', 'intento__estudiante')
            .annotate(acierto=acierto_expr)
        )

    @action(detail=False, methods=['get'], url_path='resumen')
    def resumen(self, request):
        respuestas = self._base_respuestas(request)

        promedio_general = respuestas.aggregate(
            promedio=Avg('acierto'),
            total=Count('id'),
        )

        por_dificultad = (
            respuestas.values('pregunta__nivel_dificultad')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('pregunta__nivel_dificultad')
        )

        por_competencia = (
            respuestas.values('pregunta__competencia__nombre')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('pregunta__competencia__nombre')
        )

        por_categoria = (
            respuestas.values('pregunta__categoria__nombre')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('pregunta__categoria__nombre')
        )

        por_genero = (
            respuestas.values('intento__estudiante__genero')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('intento__estudiante__genero')
        )

        por_semestre = (
            respuestas.values('intento__estudiante__semestre_actual')
            .annotate(promedio=Avg('acierto'), total=Count('id'))
            .order_by('intento__estudiante__semestre_actual')
        )

        def _map_bucket(rows, key):
            data = []
            for row in rows:
                etiqueta = row.get(key)
                promedio = float(row.get('promedio') or 0.0)
                data.append(
                    {
                        'label': etiqueta if etiqueta not in [None, ''] else 'Sin dato',
                        'promedio': round(promedio * 300, 2),
                        'total': int(row.get('total') or 0),
                    }
                )
            return data

        return Response(
            {
                'promedio_general_prueba': round(float(promedio_general.get('promedio') or 0.0) * 300, 2),
                'total_respuestas': int(promedio_general.get('total') or 0),
                'promedio_por_dificultad': _map_bucket(por_dificultad, 'pregunta__nivel_dificultad'),
                'promedio_por_competencia': _map_bucket(por_competencia, 'pregunta__competencia__nombre'),
                'promedio_por_categoria': _map_bucket(por_categoria, 'pregunta__categoria__nombre'),
                'desempeno_por_genero': _map_bucket(por_genero, 'intento__estudiante__genero'),
                'desempeno_por_semestre': _map_bucket(por_semestre, 'intento__estudiante__semestre_actual'),
            },
            status=status.HTTP_200_OK,
        )
